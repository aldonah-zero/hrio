import uvicorn
import os, json
import time as time_module
import logging
from fastapi import Depends, FastAPI, HTTPException, Request, status, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
from pydantic_classes import *
from sql_alchemy import *
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
import resend
import os

resend.api_key = os.getenv("RESEND_API_KEY")
############################################
#
#   Initialize the database
#
############################################

def init_db():
    DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./data/Class_Diagram.db")

    # ensure sqlite folder exists
    if DATABASE_URL.startswith("sqlite"):
        os.makedirs("data", exist_ok=True)
        engine = create_engine(
            DATABASE_URL,
            connect_args={"check_same_thread": False}
        )
    else:
        engine = create_engine(
            DATABASE_URL,
            pool_size=10,
            max_overflow=20,
            pool_pre_ping=True
        )

    SessionLocal = sessionmaker(
        autocommit=False,
        autoflush=False,
        bind=engine
    )

    Base.metadata.create_all(bind=engine)

    return SessionLocal
app = FastAPI(
    title="Class_Diagram API",
    description="Auto-generated REST API with full CRUD operations, relationship management, and advanced features",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc",
    openapi_tags=[
        {"name": "System", "description": "System health and statistics"},
        {"name": "Cena", "description": "Operations for Cena entities"},
        {"name": "Cena Relationships", "description": "Manage Cena relationships"},
        {"name": "SesijaGrupa", "description": "Operations for SesijaGrupa entities"},
        {"name": "SesijaGrupa Relationships", "description": "Manage SesijaGrupa relationships"},
        {"name": "SesijaKlijent", "description": "Operations for SesijaKlijent entities"},
        {"name": "SesijaKlijent Relationships", "description": "Manage SesijaKlijent relationships"},
        {"name": "Sesija", "description": "Operations for Sesija entities"},
        {"name": "Sesija Relationships", "description": "Manage Sesija relationships"},
        {"name": "Grupa", "description": "Operations for Grupa entities"},
        {"name": "Grupa Relationships", "description": "Manage Grupa relationships"},
        {"name": "GrupaKlijent", "description": "Operations for GrupaKlijent (members of groups)"},
        {"name": "Klijent", "description": "Operations for Klijent entities"},
        {"name": "Klijent Relationships", "description": "Manage Klijent relationships"},
    ]
)


from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

scheduler = BackgroundScheduler(timezone="Europe/Belgrade")

@app.on_event("startup")
def start_scheduler():
    # Run every hour at minute 0 (e.g. 09:00, 10:00, 11:00…)
    scheduler.add_job(
        send_daily_reminders,
        trigger=CronTrigger(minute=0),
        id="daily_session_reminders",
        replace_existing=True,
        max_instances=1,
    )
    scheduler.start()
    logger.info("Scheduler started — reminders will be sent hourly")

@app.on_event("shutdown")
def shutdown_scheduler():
    scheduler.shutdown()
    logger.info("Scheduler stopped")

def format_datetime(dt):
    return dt.strftime("%d.%m.%Y %H:%M")

def format_date_long(dt):
    days = ["ponedeljak", "utorak", "sreda", "četvrtak", "petak", "subota", "nedelja"]
    months = ["januar", "februar", "mart", "april", "maj", "jun", "jul", "avgust", "septembar", "oktobar", "novembar", "decembar"]
    return f"{days[dt.weekday()]}, {dt.day}. {months[dt.month - 1]} {dt.year}."

def format_time(dt):
    return dt.strftime("%H:%M")


def send_session_email(action, client_name, pocetak, kraj, cena, client_email=None):

    config = {
        "created": {
            "title": "Sesija zakazana",
            "greeting": "Vaša sesija je potvrđena!",
            "color": "#4f46e5",
            "icon": "✅"
        },
        "updated": {
            "title": "Sesija izmenjena",
            "greeting": "Vaša sesija je ažurirana.",
            "color": "#f59e0b",
            "icon": "✏️"
        },
        "deleted": {
            "title": "Sesija otkazana",
            "greeting": "Vaša sesija je otkazana.",
            "color": "#ef4444",
            "icon": "🗑️"
        }
    }

    c = config[action]
    app_url = os.getenv("APP_URL", "https://hrio-frontend-5c8704.onrender.com/")

    html = f"""
<div style="background:#f2f2f7;padding:32px 16px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,sans-serif;color:#1a1a1a;">
<div style="max-width:520px;margin:auto;">

<div style="background:#fff;border-radius:16px;padding:28px 24px 24px;margin-bottom:8px;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
<div style="font-size:22px;margin-bottom:8px;">{c['icon']} {c['title']}</div>
<div style="font-size:15px;color:#1a1a1a;margin-bottom:4px;">Zdravo, <strong>Maja</strong>.</div>
<div style="font-size:15px;color:#1a1a1a;margin-bottom:4px;">{c['greeting']}</div>
<div style="font-size:14px;color:#555;line-height:1.5;">Evo nekoliko detalja koje biste trebali znati o svojoj sesiji:</div>
</div>

<div style="background:#fff;border-radius:16px;padding:24px;margin-bottom:8px;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
<div style="border:1.5px dashed #d1d5db;border-radius:12px;padding:20px;">

<div style="font-size:15px;color:#333;margin-bottom:14px;">📅 <strong>{format_date_long(pocetak)}</strong></div>

<div style="font-size:15px;font-weight:700;color:#111;margin-bottom:2px;">Individualna sesija</div>
<div style="font-size:15px;font-weight:600;color:#111;margin-bottom:4px;">{format_time(pocetak)} – {format_time(kraj)}</div>
<div style="font-size:14px;color:#555;margin-bottom:16px;">{client_name} · {cena:,.2f} RSD</div>

<div style="border-top:1.5px dashed #d1d5db;margin:0 0 16px;"></div>

<table cellpadding="0" cellspacing="0" width="100%">
<tr><td style="font-size:14px;color:#333;padding:3px 0;"><strong>Ukupno:</strong></td><td style="font-size:14px;color:#333;padding:3px 0;text-align:right;"><strong>{cena:,.2f} RSD</strong></td></tr>
<tr><td style="font-size:14px;color:#333;padding:3px 0;"><strong>Status:</strong></td><td style="font-size:14px;color:#333;padding:3px 0;text-align:right;">{c['title']}</td></tr>
</table>

</div>
</div>

<div style="background:#fff;border-radius:16px;padding:20px 24px;margin-bottom:8px;text-align:center;box-shadow:0 1px 6px rgba(0,0,0,0.04);">

<div style="font-size:13px;color:#777;margin-bottom:14px;line-height:1.5;"><strong>Pravila otkazivanja i kašnjenja</strong><br>Vaša sesija može biti otkazana do <strong>24 sata</strong> pre termina. Ako kasnite više od <strong>10 minuta</strong>, sesija će se smatrati propuštenom.</div>

<a href="{app_url}" style="display:inline-block;padding:14px 36px;background:#0f172a;color:#ffffff;text-decoration:none;border-radius:10px;font-size:15px;font-weight:600;">Otvori aplikaciju</a>
</div>

<div style="text-align:center;font-size:13px;color:#9ca3af;margin-top:14px;line-height:1.5;">
Hvala vam što ste odabrali <strong style="color:#6b7280;">PsihApp</strong>.
</div>

</div>
</div>
"""

    # Email za psihologa (tvoj)
    resend.Emails.send({
        "from": "Hrio <noreply@hrioapp.com>",
        "to": ["igorpavlov106@gmail.com"],
        "subject": f"{c['icon']} {c['title']} - {client_name}",
        "html": html
    })

    # Email za klijenta (samo ako ima email)
    if client_email:
        client_html = f"""
    <div style="background:#f2f2f7;padding:32px 16px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,sans-serif;color:#1a1a1a;">
    <div style="max-width:520px;margin:auto;">

    <div style="background:#fff;border-radius:16px;padding:28px 24px 24px;margin-bottom:8px;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
    <div style="font-size:22px;margin-bottom:8px;">✅ Potvrda zakazane sesije</div>
    <div style="font-size:15px;color:#1a1a1a;margin-bottom:4px;">Poštovani/a <strong>{client_name}</strong>,</div>
    <div style="font-size:15px;color:#1a1a1a;margin-bottom:4px;">Vaša sesija je uspešno zakazana.</div>
    </div>

    <div style="background:#fff;border-radius:16px;padding:24px;margin-bottom:8px;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
    <div style="border:1.5px dashed #d1d5db;border-radius:12px;padding:20px;">

    <div style="font-size:15px;color:#333;margin-bottom:14px;">📅 <strong>{format_date_long(pocetak)}</strong></div>
    <div style="font-size:15px;font-weight:600;color:#111;margin-bottom:4px;">{format_time(pocetak)} – {format_time(kraj)}</div>

    <div style="border-top:1.5px dashed #d1d5db;margin:16px 0;"></div>

    <div style="font-size:14px;color:#555;line-height:1.6;">
    Molimo Vas da dođete <strong>5 minuta</strong> pre zakazanog termina.
    </div>

    </div>
    </div>

    <div style="background:#fff;border-radius:16px;padding:20px 24px;margin-bottom:8px;text-align:center;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
    <div style="font-size:13px;color:#777;line-height:1.5;"><strong>Pravila otkazivanja</strong><br>Sesija se može otkazati najkasnije <strong>24 sata</strong> pre termina.<br>Kašnjenje duže od <strong>10 minuta</strong> smatra se propuštenom sesijom.</div>
    </div>

    <div style="text-align:center;font-size:13px;color:#9ca3af;margin-top:14px;line-height:1.5;">
    Hvala vam na poverenju. <strong style="color:#6b7280;">PsihApp</strong>
    </div>

    </div>
    </div>
    """
        resend.Emails.send({
            "from": "Hrio <noreply@hrioapp.com>",
            "to": [client_email],
            "subject": f"✅ Potvrda sesije - {format_date_long(pocetak)}",
            "html": client_html
        })


def send_session_email_to_group(action, grupa_naziv, pocetak, kraj, cena, client_emails):
    """Send session email to all members of a group who have email addresses."""
    config = {
        "created": {
            "title": "Grupna sesija zakazana",
            "greeting": "Vaša grupna sesija je potvrđena!",
            "icon": "✅"
        },
        "updated": {
            "title": "Grupna sesija izmenjena",
            "greeting": "Vaša grupna sesija je ažurirana.",
            "icon": "✏️"
        },
        "deleted": {
            "title": "Grupna sesija otkazana",
            "greeting": "Vaša grupna sesija je otkazana.",
            "icon": "🗑️"
        }
    }

    c = config[action]
    app_url = os.getenv("APP_URL", "https://hrio-frontend-5c8704.onrender.com/")

    psiholog_html = f"""
<div style="background:#f2f2f7;padding:32px 16px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,sans-serif;color:#1a1a1a;">
<div style="max-width:520px;margin:auto;">

<div style="background:#fff;border-radius:16px;padding:28px 24px 24px;margin-bottom:8px;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
<div style="font-size:22px;margin-bottom:8px;">{c['icon']} {c['title']}</div>
<div style="font-size:15px;color:#1a1a1a;margin-bottom:4px;">Zdravo, <strong>Maja</strong>.</div>
<div style="font-size:15px;color:#1a1a1a;margin-bottom:4px;">{c['greeting']}</div>
</div>

<div style="background:#fff;border-radius:16px;padding:24px;margin-bottom:8px;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
<div style="border:1.5px dashed #d1d5db;border-radius:12px;padding:20px;">

<div style="font-size:15px;color:#333;margin-bottom:14px;">📅 <strong>{format_date_long(pocetak)}</strong></div>

<div style="font-size:15px;font-weight:700;color:#111;margin-bottom:2px;">Grupna sesija — {grupa_naziv}</div>
<div style="font-size:15px;font-weight:600;color:#111;margin-bottom:4px;">{format_time(pocetak)} – {format_time(kraj)}</div>
<div style="font-size:14px;color:#555;margin-bottom:16px;">{cena:,.2f} RSD</div>

<div style="border-top:1.5px dashed #d1d5db;margin:0 0 16px;"></div>

<table cellpadding="0" cellspacing="0" width="100%">
<tr><td style="font-size:14px;color:#333;padding:3px 0;"><strong>Ukupno:</strong></td><td style="font-size:14px;color:#333;padding:3px 0;text-align:right;"><strong>{cena:,.2f} RSD</strong></td></tr>
<tr><td style="font-size:14px;color:#333;padding:3px 0;"><strong>Status:</strong></td><td style="font-size:14px;color:#333;padding:3px 0;text-align:right;">{c['title']}</td></tr>
</table>

</div>
</div>

<div style="text-align:center;font-size:13px;color:#9ca3af;margin-top:14px;line-height:1.5;">
<strong style="color:#6b7280;">PsihApp</strong>
</div>

</div>
</div>
"""

    resend.Emails.send({
        "from": "Hrio <noreply@hrioapp.com>",
        "to": ["igorpavlov106@gmail.com"],
        "subject": f"{c['icon']} {c['title']} - {grupa_naziv}",
        "html": psiholog_html
    })

    for member in client_emails:
        if not member.get("email"):
            continue

        member_name = member.get("name", "Poštovani/a")

        member_html = f"""
    <div style="background:#f2f2f7;padding:32px 16px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,sans-serif;color:#1a1a1a;">
    <div style="max-width:520px;margin:auto;">

    <div style="background:#fff;border-radius:16px;padding:28px 24px 24px;margin-bottom:8px;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
    <div style="font-size:22px;margin-bottom:8px;">{c['icon']} {c['title']}</div>
    <div style="font-size:15px;color:#1a1a1a;margin-bottom:4px;">Poštovani/a <strong>{member_name}</strong>,</div>
    <div style="font-size:15px;color:#1a1a1a;margin-bottom:4px;">{c['greeting']}</div>
    </div>

    <div style="background:#fff;border-radius:16px;padding:24px;margin-bottom:8px;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
    <div style="border:1.5px dashed #d1d5db;border-radius:12px;padding:20px;">

    <div style="font-size:15px;color:#333;margin-bottom:14px;">📅 <strong>{format_date_long(pocetak)}</strong></div>
    <div style="font-size:15px;font-weight:700;color:#111;margin-bottom:2px;">Grupna sesija — {grupa_naziv}</div>
    <div style="font-size:15px;font-weight:600;color:#111;margin-bottom:4px;">{format_time(pocetak)} – {format_time(kraj)}</div>

    <div style="border-top:1.5px dashed #d1d5db;margin:16px 0;"></div>

    <div style="font-size:14px;color:#555;line-height:1.6;">
    Molimo Vas da dođete <strong>5 minuta</strong> pre zakazanog termina.
    </div>

    </div>
    </div>

    <div style="background:#fff;border-radius:16px;padding:20px 24px;margin-bottom:8px;text-align:center;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
    <div style="font-size:13px;color:#777;line-height:1.5;"><strong>Pravila otkazivanja</strong><br>Sesija se može otkazati najkasnije <strong>24 sata</strong> pre termina.<br>Kašnjenje duže od <strong>10 minuta</strong> smatra se propuštenom sesijom.</div>
    </div>

    <div style="text-align:center;font-size:13px;color:#9ca3af;margin-top:14px;line-height:1.5;">
    Hvala vam na poverenju. <strong style="color:#6b7280;">PsihApp</strong>
    </div>

    </div>
    </div>
    """
        resend.Emails.send({
            "from": "Hrio <noreply@hrioapp.com>",
            "to": [member["email"]],
            "subject": f"✅ Potvrda grupne sesije - {format_date_long(pocetak)}",
            "html": member_html
        })


# Enable CORS for all origins (for development)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

############################################
#
#   Middleware
#
############################################

@app.middleware("http")
async def log_requests(request: Request, call_next):
    logger.info(f"Incoming request: {request.method} {request.url.path}")
    response = await call_next(request)
    logger.info(f"Response status: {response.status_code}")
    return response

@app.middleware("http")
async def add_process_time_header(request: Request, call_next):
    start_time = time_module.time()
    response = await call_next(request)
    process_time = time_module.time() - start_time
    response.headers["X-Process-Time"] = str(process_time)
    return response

############################################
#
#   Exception Handlers
#
############################################

@app.exception_handler(ValueError)
async def value_error_handler(request: Request, exc: ValueError):
    return JSONResponse(
        status_code=status.HTTP_400_BAD_REQUEST,
        content={"error": "Bad Request", "message": str(exc), "detail": "Invalid input data provided"}
    )

@app.exception_handler(IntegrityError)
async def integrity_error_handler(request: Request, exc: IntegrityError):
    logger.error(f"Database integrity error: {exc}")
    error_detail = str(exc.orig) if hasattr(exc, 'orig') else str(exc)
    return JSONResponse(
        status_code=status.HTTP_409_CONFLICT,
        content={"error": "Conflict", "message": "Data conflict occurred", "detail": error_detail}
    )

@app.exception_handler(SQLAlchemyError)
async def sqlalchemy_error_handler(request: Request, exc: SQLAlchemyError):
    logger.error(f"Database error: {exc}")
    return JSONResponse(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        content={"error": "Internal Server Error", "message": "Database operation failed", "detail": "An internal database error occurred"}
    )

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    return JSONResponse(
        status_code=exc.status_code,
        content={"error": exc.detail if isinstance(exc.detail, str) else "HTTP Error", "message": exc.detail, "detail": f"HTTP {exc.status_code} error occurred"}
    )

# Initialize database session
SessionLocal = init_db()
def get_db():
    db = SessionLocal()
    try:
        yield db
    except Exception:
        db.rollback()
        logger.error("Database session rollback due to exception")
        raise
    finally:
        db.close()

############################################
#
#   Global API endpoints
#
############################################

def send_reminder_email(client_name, client_email, pocetak, kraj, is_group=False, grupa_naziv=None):
    """Send a 24h-before reminder email to a client."""
    if not client_email:
        return

    app_url = os.getenv("APP_URL", "https://hrio-frontend-5c8704.onrender.com/")

    session_type_label = f"Grupna sesija — {grupa_naziv}" if is_group else "Individualna sesija"

    html = f"""
<div style="background:#f2f2f7;padding:32px 16px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,sans-serif;color:#1a1a1a;">
<div style="max-width:520px;margin:auto;">

<div style="background:#fff;border-radius:16px;padding:28px 24px 24px;margin-bottom:8px;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
<div style="font-size:22px;margin-bottom:8px;">⏰ Podsetnik za sesiju sutra</div>
<div style="font-size:15px;color:#1a1a1a;margin-bottom:4px;">Poštovani/a <strong>{client_name}</strong>,</div>
<div style="font-size:15px;color:#1a1a1a;margin-bottom:4px;">Podsećamo Vas da imate zakazanu sesiju <strong>sutra</strong>.</div>
</div>

<div style="background:#fff;border-radius:16px;padding:24px;margin-bottom:8px;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
<div style="border:1.5px dashed #d1d5db;border-radius:12px;padding:20px;">

<div style="font-size:15px;color:#333;margin-bottom:14px;">📅 <strong>{format_date_long(pocetak)}</strong></div>
<div style="font-size:15px;font-weight:700;color:#111;margin-bottom:2px;">{session_type_label}</div>
<div style="font-size:15px;font-weight:600;color:#111;margin-bottom:4px;">{format_time(pocetak)} – {format_time(kraj)}</div>

<div style="border-top:1.5px dashed #d1d5db;margin:16px 0;"></div>

<div style="font-size:14px;color:#555;line-height:1.6;">
Molimo Vas da dođete <strong>5 minuta</strong> pre zakazanog termina.
</div>

</div>
</div>

<div style="background:#fff;border-radius:16px;padding:20px 24px;margin-bottom:8px;text-align:center;box-shadow:0 1px 6px rgba(0,0,0,0.04);">
<div style="font-size:13px;color:#777;line-height:1.5;"><strong>Pravila otkazivanja</strong><br>Sesija se može otkazati najkasnije <strong>24 sata</strong> pre termina.<br>Kašnjenje duže od <strong>10 minuta</strong> smatra se propuštenom sesijom.</div>
</div>

<div style="text-align:center;font-size:13px;color:#9ca3af;margin-top:14px;line-height:1.5;">
Vidimo se sutra! <strong style="color:#6b7280;">PsihApp</strong>
</div>

</div>
</div>
"""

    try:
        resend.Emails.send({
            "from": "Hrio <noreply@hrioapp.com>",
            "to": [client_email],
            "subject": f"⏰ Podsetnik: sesija sutra u {format_time(pocetak)}",
            "html": html
        })
        logger.info(f"Reminder sent to {client_email} for session at {pocetak}")
    except Exception as e:
        logger.error(f"Failed to send reminder to {client_email}: {e}")


def send_daily_reminders():
    """
    Scans all sessions happening ~24h from now (within a 1-hour window)
    and sends a reminder email to each client.
    Runs every hour — the 1-hour window prevents missing or duplicating.
    """
    from datetime import datetime, timedelta

    db = SessionLocal()
    try:
        now = datetime.now()
        window_start = now + timedelta(hours=23, minutes=30)
        window_end = now + timedelta(hours=24, minutes=30)

        logger.info(f"Reminder job: scanning sessions between {window_start} and {window_end}")

        upcoming_sessions = db.query(Sesija).filter(
            Sesija.pocetak >= window_start,
            Sesija.pocetak < window_end,
            Sesija.status == "zakazano"
        ).all()

        logger.info(f"Reminder job: found {len(upcoming_sessions)} session(s) needing reminders")

        for sesija in upcoming_sessions:
            try:
                sk = db.query(SesijaKlijent).filter(SesijaKlijent.sesija_id == sesija.id).first()
                if sk and sk.klijent_id:
                    klijent = db.query(Klijent).filter(Klijent.id == sk.klijent_id).first()
                    if klijent and klijent.email:
                        send_reminder_email(
                            client_name=f"{klijent.ime} {klijent.prezime}",
                            client_email=klijent.email,
                            pocetak=sesija.pocetak,
                            kraj=sesija.kraj,
                            is_group=False
                        )
                    continue

                sg = db.query(SesijaGrupa).filter(SesijaGrupa.sesija_1_id == sesija.id).first()
                if sg and sg.grupa_id:
                    grupa = db.query(Grupa).filter(Grupa.id == sg.grupa_id).first()
                    if grupa:
                        members = db.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == grupa.id).all()
                        for gk in members:
                            klijent = db.query(Klijent).filter(Klijent.id == gk.klijent_id).first()
                            if klijent and klijent.email:
                                send_reminder_email(
                                    client_name=f"{klijent.ime} {klijent.prezime}",
                                    client_email=klijent.email,
                                    pocetak=sesija.pocetak,
                                    kraj=sesija.kraj,
                                    is_group=True,
                                    grupa_naziv=grupa.naziv
                                )
            except Exception as e:
                logger.error(f"Error processing reminder for session {sesija.id}: {e}")

    except Exception as e:
        logger.error(f"Reminder job failed: {e}")
    finally:
        db.close()
@app.get("/", tags=["System"])
def root():
    return {"name": "Class_Diagram API", "version": "1.0.0", "status": "running"}

@app.get("/health", tags=["System"])
def health_check():
    from datetime import datetime
    return {"status": "healthy", "timestamp": datetime.now().isoformat(), "database": "connected"}

@app.get("/statistics", tags=["System"])
def get_statistics(database: Session = Depends(get_db)):
    stats = {}
    stats["cena_count"] = database.query(Cena).count()
    stats["sesijagrupa_count"] = database.query(SesijaGrupa).count()
    stats["sesijaklijent_count"] = database.query(SesijaKlijent).count()
    stats["sesija_count"] = database.query(Sesija).count()
    stats["grupa_count"] = database.query(Grupa).count()
    stats["klijent_count"] = database.query(Klijent).count()
    stats["grupaklijent_count"] = database.query(GrupaKlijent).count()
    stats["total_entities"] = sum(stats.values())
    return stats

############################################
#
#   BESSER Action Language standard lib
#
############################################

async def BAL_size(sequence:list) -> int:
    return len(sequence)
async def BAL_is_empty(sequence:list) -> bool:
    return len(sequence) == 0
async def BAL_add(sequence:list, elem) -> None:
    sequence.append(elem)
async def BAL_remove(sequence:list, elem) -> None:
    sequence.remove(elem)
async def BAL_contains(sequence:list, elem) -> bool:
    return elem in sequence
async def BAL_filter(sequence:list, predicate) -> list:
    return [elem for elem in sequence if predicate(elem)]
async def BAL_forall(sequence:list, predicate) -> bool:
    for elem in sequence:
        if not predicate(elem):
            return False
    return True
async def BAL_exists(sequence:list, predicate) -> bool:
    for elem in sequence:
        if predicate(elem):
            return True
    return False
async def BAL_one(sequence:list, predicate) -> bool:
    found = False
    for elem in sequence:
        if predicate(elem):
            if found:
                return False
            found = True
    return found
async def BAL_is_unique(sequence:list, mapping) -> bool:
    mapped = [mapping(elem) for elem in sequence]
    return len(set(mapped)) == len(mapped)
async def BAL_map(sequence:list, mapping) -> list:
    return [mapping(elem) for elem in sequence]
async def BAL_reduce(sequence:list, reduce_fn, aggregator) -> any:
    for elem in sequence:
        aggregator = reduce_fn(aggregator, elem)
    return aggregator


############################################
#
#   Cena functions
#
############################################

@app.get("/cena/", response_model=None, tags=["Cena"])
def get_all_cena(detailed: bool = False, database: Session = Depends(get_db)) -> list:
    from sqlalchemy.orm import joinedload
    if detailed:
        query = database.query(Cena)
        query = query.options(joinedload(Cena.sesija_2))
        query = query.options(joinedload(Cena.klijent_1))
        cena_list = query.all()
        result = []
        for cena_item in cena_list:
            item_dict = cena_item.__dict__.copy()
            item_dict.pop('_sa_instance_state', None)
            if cena_item.sesija_2:
                related_dict = cena_item.sesija_2.__dict__.copy()
                related_dict.pop('_sa_instance_state', None)
                item_dict['sesija_2'] = related_dict
            else:
                item_dict['sesija_2'] = None
            if cena_item.klijent_1:
                related_dict = cena_item.klijent_1.__dict__.copy()
                related_dict.pop('_sa_instance_state', None)
                item_dict['klijent_1'] = related_dict
            else:
                item_dict['klijent_1'] = None
            result.append(item_dict)
        return result
    else:
        return database.query(Cena).all()

@app.get("/cena/count/", response_model=None, tags=["Cena"])
def get_count_cena(database: Session = Depends(get_db)) -> dict:
    return {"count": database.query(Cena).count()}

@app.get("/cena/paginated/", response_model=None, tags=["Cena"])
def get_paginated_cena(skip: int = 0, limit: int = 100, detailed: bool = False, database: Session = Depends(get_db)) -> dict:
    total = database.query(Cena).count()
    cena_list = database.query(Cena).offset(skip).limit(limit).all()
    return {"total": total, "skip": skip, "limit": limit, "data": cena_list}

@app.get("/cena/search/", response_model=None, tags=["Cena"])
def search_cena(database: Session = Depends(get_db)) -> list:
    return database.query(Cena).all()

@app.get("/cena/{cena_id}/", response_model=None, tags=["Cena"])
async def get_cena(cena_id: int, database: Session = Depends(get_db)) -> Cena:
    db_cena = database.query(Cena).filter(Cena.id == cena_id).first()
    if db_cena is None:
        raise HTTPException(status_code=404, detail="Cena not found")
    return {"cena": db_cena}

@app.post("/cena/", response_model=None, tags=["Cena"])
async def create_cena(cena_data: CenaCreate, database: Session = Depends(get_db)) -> Cena:
    if cena_data.sesija_2 is not None:
        db_sesija_2 = database.query(Sesija).filter(Sesija.id == cena_data.sesija_2).first()
        if not db_sesija_2:
            raise HTTPException(status_code=400, detail="Sesija not found")
    else:
        raise HTTPException(status_code=400, detail="Sesija ID is required")
    if cena_data.klijent_1 is not None:
        db_klijent_1 = database.query(Klijent).filter(Klijent.id == cena_data.klijent_1).first()
        if not db_klijent_1:
            raise HTTPException(status_code=400, detail="Klijent not found")
    else:
        raise HTTPException(status_code=400, detail="Klijent ID is required")
    cena_kwargs = dict(cena=cena_data.cena, status=cena_data.status, nacin_placanja=cena_data.nacin_placanja, datum_uplate=cena_data.datum_uplate, sesija_2_id=cena_data.sesija_2, klijent_1_id=cena_data.klijent_1)
    if cena_data.id is not None:
        cena_kwargs["id"] = cena_data.id
    db_cena = Cena(**cena_kwargs)
    database.add(db_cena)
    database.commit()
    database.refresh(db_cena)
    return db_cena

@app.post("/cena/bulk/", response_model=None, tags=["Cena"])
async def bulk_create_cena(items: list[CenaCreate], database: Session = Depends(get_db)) -> dict:
    created_items = []
    errors = []
    for idx, item_data in enumerate(items):
        try:
            if not item_data.sesija_2:
                raise ValueError("Sesija ID is required")
            if not item_data.klijent_1:
                raise ValueError("Klijent ID is required")
            db_cena = Cena(cena=item_data.cena, id=item_data.id, status=item_data.status, nacin_placanja=item_data.nacin_placanja, datum_uplate=item_data.datum_uplate, sesija_2_id=item_data.sesija_2, klijent_1_id=item_data.klijent_1)
            database.add(db_cena)
            database.flush()
            created_items.append(db_cena.id)
        except Exception as e:
            errors.append({"index": idx, "error": str(e)})
    if errors:
        database.rollback()
        raise HTTPException(status_code=400, detail={"message": "Bulk creation failed", "errors": errors})
    database.commit()
    return {"created_count": len(created_items), "created_ids": created_items, "message": f"Successfully created {len(created_items)} Cena entities"}

@app.delete("/cena/bulk/", response_model=None, tags=["Cena"])
async def bulk_delete_cena(ids: list[int], database: Session = Depends(get_db)) -> dict:
    deleted_count = 0
    not_found = []
    for item_id in ids:
        db_cena = database.query(Cena).filter(Cena.id == item_id).first()
        if db_cena:
            database.delete(db_cena)
            deleted_count += 1
        else:
            not_found.append(item_id)
    database.commit()
    return {"deleted_count": deleted_count, "not_found": not_found, "message": f"Successfully deleted {deleted_count} Cena entities"}

@app.put("/cena/{cena_id}/", response_model=None, tags=["Cena"])
async def update_cena(cena_id: int, cena_data: CenaCreate, database: Session = Depends(get_db)) -> Cena:
    db_cena = database.query(Cena).filter(Cena.id == cena_id).first()
    if db_cena is None:
        raise HTTPException(status_code=404, detail="Cena not found")
    setattr(db_cena, 'cena', cena_data.cena)
    setattr(db_cena, 'id', cena_data.id)
    setattr(db_cena, 'status', cena_data.status)
    setattr(db_cena, 'nacin_placanja', cena_data.nacin_placanja)
    setattr(db_cena, 'datum_uplate', cena_data.datum_uplate)
    if cena_data.sesija_2 is not None:
        db_sesija_2 = database.query(Sesija).filter(Sesija.id == cena_data.sesija_2).first()
        if not db_sesija_2:
            raise HTTPException(status_code=400, detail="Sesija not found")
        setattr(db_cena, 'sesija_2_id', cena_data.sesija_2)
    if cena_data.klijent_1 is not None:
        db_klijent_1 = database.query(Klijent).filter(Klijent.id == cena_data.klijent_1).first()
        if not db_klijent_1:
            raise HTTPException(status_code=400, detail="Klijent not found")
        setattr(db_cena, 'klijent_1_id', cena_data.klijent_1)
    database.commit()
    database.refresh(db_cena)
    return db_cena

@app.delete("/cena/{cena_id}/", response_model=None, tags=["Cena"])
async def delete_cena(cena_id: int, database: Session = Depends(get_db)):
    db_cena = database.query(Cena).filter(Cena.id == cena_id).first()
    if db_cena is None:
        raise HTTPException(status_code=404, detail="Cena not found")
    database.delete(db_cena)
    database.commit()
    return {"message": "Deleted", "id": cena_id}


############################################
#
#   SesijaGrupa functions
#
############################################

@app.get("/sesijagrupa/", response_model=None, tags=["SesijaGrupa"])
def get_all_sesijagrupa(detailed: bool = False, database: Session = Depends(get_db)) -> list:
    from sqlalchemy.orm import joinedload
    if detailed:
        query = database.query(SesijaGrupa)
        query = query.options(joinedload(SesijaGrupa.grupa))
        query = query.options(joinedload(SesijaGrupa.sesija_1))
        sesijagrupa_list = query.all()
        result = []
        for sesijagrupa_item in sesijagrupa_list:
            item_dict = sesijagrupa_item.__dict__.copy()
            item_dict.pop('_sa_instance_state', None)
            if sesijagrupa_item.grupa:
                related_dict = sesijagrupa_item.grupa.__dict__.copy()
                related_dict.pop('_sa_instance_state', None)
                item_dict['grupa'] = related_dict
            else:
                item_dict['grupa'] = None
            if sesijagrupa_item.sesija_1:
                related_dict = sesijagrupa_item.sesija_1.__dict__.copy()
                related_dict.pop('_sa_instance_state', None)
                item_dict['sesija_1'] = related_dict
            else:
                item_dict['sesija_1'] = None
            result.append(item_dict)
        return result
    else:
        return database.query(SesijaGrupa).all()

@app.get("/sesijagrupa/count/", response_model=None, tags=["SesijaGrupa"])
def get_count_sesijagrupa(database: Session = Depends(get_db)) -> dict:
    return {"count": database.query(SesijaGrupa).count()}

@app.get("/sesijagrupa/paginated/", response_model=None, tags=["SesijaGrupa"])
def get_paginated_sesijagrupa(skip: int = 0, limit: int = 100, detailed: bool = False, database: Session = Depends(get_db)) -> dict:
    total = database.query(SesijaGrupa).count()
    sesijagrupa_list = database.query(SesijaGrupa).offset(skip).limit(limit).all()
    return {"total": total, "skip": skip, "limit": limit, "data": sesijagrupa_list}

@app.get("/sesijagrupa/search/", response_model=None, tags=["SesijaGrupa"])
def search_sesijagrupa(database: Session = Depends(get_db)) -> list:
    return database.query(SesijaGrupa).all()

@app.get("/sesijagrupa/{sesijagrupa_id}/", response_model=None, tags=["SesijaGrupa"])
async def get_sesijagrupa(sesijagrupa_id: int, database: Session = Depends(get_db)) -> SesijaGrupa:
    db_sesijagrupa = database.query(SesijaGrupa).filter(SesijaGrupa.id == sesijagrupa_id).first()
    if db_sesijagrupa is None:
        raise HTTPException(status_code=404, detail="SesijaGrupa not found")
    return {"sesijagrupa": db_sesijagrupa}

@app.post("/sesijagrupa/", response_model=None, tags=["SesijaGrupa"])
async def create_sesijagrupa(sesijagrupa_data: SesijaGrupaCreate, database: Session = Depends(get_db)) -> SesijaGrupa:
    if sesijagrupa_data.grupa is not None:
        db_grupa = database.query(Grupa).filter(Grupa.id == sesijagrupa_data.grupa).first()
        if not db_grupa:
            raise HTTPException(status_code=400, detail="Grupa not found")
    else:
        raise HTTPException(status_code=400, detail="Grupa ID is required")
    if sesijagrupa_data.sesija_1 is not None:
        db_sesija_1 = database.query(Sesija).filter(Sesija.id == sesijagrupa_data.sesija_1).first()
        if not db_sesija_1:
            raise HTTPException(status_code=400, detail="Sesija not found")
    else:
        raise HTTPException(status_code=400, detail="Sesija ID is required")
    sg_kwargs = dict(grupa_id=sesijagrupa_data.grupa, sesija_1_id=sesijagrupa_data.sesija_1)
    if sesijagrupa_data.id is not None:
        sg_kwargs["id"] = sesijagrupa_data.id
    db_sesijagrupa = SesijaGrupa(**sg_kwargs)
    database.add(db_sesijagrupa)
    database.commit()
    database.refresh(db_sesijagrupa)
    return db_sesijagrupa

@app.post("/sesijagrupa/bulk/", response_model=None, tags=["SesijaGrupa"])
async def bulk_create_sesijagrupa(items: list[SesijaGrupaCreate], database: Session = Depends(get_db)) -> dict:
    created_items = []
    errors = []
    for idx, item_data in enumerate(items):
        try:
            if not item_data.grupa:
                raise ValueError("Grupa ID is required")
            if not item_data.sesija_1:
                raise ValueError("Sesija ID is required")
            db_sesijagrupa = SesijaGrupa(id=item_data.id, grupa_id=item_data.grupa, sesija_1_id=item_data.sesija_1)
            database.add(db_sesijagrupa)
            database.flush()
            created_items.append(db_sesijagrupa.id)
        except Exception as e:
            errors.append({"index": idx, "error": str(e)})
    if errors:
        database.rollback()
        raise HTTPException(status_code=400, detail={"message": "Bulk creation failed", "errors": errors})
    database.commit()
    return {"created_count": len(created_items), "created_ids": created_items, "message": f"Successfully created {len(created_items)} SesijaGrupa entities"}

@app.delete("/sesijagrupa/bulk/", response_model=None, tags=["SesijaGrupa"])
async def bulk_delete_sesijagrupa(ids: list[int], database: Session = Depends(get_db)) -> dict:
    deleted_count = 0
    not_found = []
    for item_id in ids:
        db_sesijagrupa = database.query(SesijaGrupa).filter(SesijaGrupa.id == item_id).first()
        if db_sesijagrupa:
            database.delete(db_sesijagrupa)
            deleted_count += 1
        else:
            not_found.append(item_id)
    database.commit()
    return {"deleted_count": deleted_count, "not_found": not_found, "message": f"Successfully deleted {deleted_count} SesijaGrupa entities"}

@app.put("/sesijagrupa/{sesijagrupa_id}/", response_model=None, tags=["SesijaGrupa"])
async def update_sesijagrupa(sesijagrupa_id: int, sesijagrupa_data: SesijaGrupaCreate, database: Session = Depends(get_db)) -> SesijaGrupa:
    db_sesijagrupa = database.query(SesijaGrupa).filter(SesijaGrupa.id == sesijagrupa_id).first()
    if db_sesijagrupa is None:
        raise HTTPException(status_code=404, detail="SesijaGrupa not found")
    setattr(db_sesijagrupa, 'id', sesijagrupa_data.id)
    if sesijagrupa_data.grupa is not None:
        db_grupa = database.query(Grupa).filter(Grupa.id == sesijagrupa_data.grupa).first()
        if not db_grupa:
            raise HTTPException(status_code=400, detail="Grupa not found")
        setattr(db_sesijagrupa, 'grupa_id', sesijagrupa_data.grupa)
    if sesijagrupa_data.sesija_1 is not None:
        db_sesija_1 = database.query(Sesija).filter(Sesija.id == sesijagrupa_data.sesija_1).first()
        if not db_sesija_1:
            raise HTTPException(status_code=400, detail="Sesija not found")
        setattr(db_sesijagrupa, 'sesija_1_id', sesijagrupa_data.sesija_1)
    database.commit()
    database.refresh(db_sesijagrupa)
    return db_sesijagrupa

@app.delete("/sesijagrupa/{sesijagrupa_id}/", response_model=None, tags=["SesijaGrupa"])
async def delete_sesijagrupa(sesijagrupa_id: int, database: Session = Depends(get_db)):
    db_sesijagrupa = database.query(SesijaGrupa).filter(SesijaGrupa.id == sesijagrupa_id).first()
    if db_sesijagrupa is None:
        raise HTTPException(status_code=404, detail="SesijaGrupa not found")
    database.delete(db_sesijagrupa)
    database.commit()
    return {"message": "Deleted", "id": sesijagrupa_id}


############################################
#
#   SesijaKlijent functions
#
############################################

@app.get("/sesijaklijent/", response_model=None, tags=["SesijaKlijent"])
def get_all_sesijaklijent(detailed: bool = False, database: Session = Depends(get_db)) -> list:
    from sqlalchemy.orm import joinedload
    if detailed:
        query = database.query(SesijaKlijent)
        query = query.options(joinedload(SesijaKlijent.klijent))
        query = query.options(joinedload(SesijaKlijent.sesija))
        sesijaklijent_list = query.all()
        result = []
        for sesijaklijent_item in sesijaklijent_list:
            item_dict = sesijaklijent_item.__dict__.copy()
            item_dict.pop('_sa_instance_state', None)
            if sesijaklijent_item.klijent:
                related_dict = sesijaklijent_item.klijent.__dict__.copy()
                related_dict.pop('_sa_instance_state', None)
                item_dict['klijent'] = related_dict
            else:
                item_dict['klijent'] = None
            if sesijaklijent_item.sesija:
                related_dict = sesijaklijent_item.sesija.__dict__.copy()
                related_dict.pop('_sa_instance_state', None)
                item_dict['sesija'] = related_dict
            else:
                item_dict['sesija'] = None
            result.append(item_dict)
        return result
    else:
        return database.query(SesijaKlijent).all()

@app.get("/sesijaklijent/count/", response_model=None, tags=["SesijaKlijent"])
def get_count_sesijaklijent(database: Session = Depends(get_db)) -> dict:
    return {"count": database.query(SesijaKlijent).count()}

@app.get("/sesijaklijent/paginated/", response_model=None, tags=["SesijaKlijent"])
def get_paginated_sesijaklijent(skip: int = 0, limit: int = 100, detailed: bool = False, database: Session = Depends(get_db)) -> dict:
    total = database.query(SesijaKlijent).count()
    sesijaklijent_list = database.query(SesijaKlijent).offset(skip).limit(limit).all()
    return {"total": total, "skip": skip, "limit": limit, "data": sesijaklijent_list}

@app.get("/sesijaklijent/search/", response_model=None, tags=["SesijaKlijent"])
def search_sesijaklijent(database: Session = Depends(get_db)) -> list:
    return database.query(SesijaKlijent).all()

@app.get("/sesijaklijent/{sesijaklijent_id}/", response_model=None, tags=["SesijaKlijent"])
async def get_sesijaklijent(sesijaklijent_id: int, database: Session = Depends(get_db)) -> SesijaKlijent:
    db_sesijaklijent = database.query(SesijaKlijent).filter(SesijaKlijent.id == sesijaklijent_id).first()
    if db_sesijaklijent is None:
        raise HTTPException(status_code=404, detail="SesijaKlijent not found")
    return {"sesijaklijent": db_sesijaklijent}

@app.post("/sesijaklijent/", response_model=None, tags=["SesijaKlijent"])
async def create_sesijaklijent(sesijaklijent_data: SesijaKlijentCreate, database: Session = Depends(get_db)) -> SesijaKlijent:
    if sesijaklijent_data.klijent is not None:
        db_klijent = database.query(Klijent).filter(Klijent.id == sesijaklijent_data.klijent).first()
        if not db_klijent:
            raise HTTPException(status_code=400, detail="Klijent not found")
    else:
        raise HTTPException(status_code=400, detail="Klijent ID is required")
    if sesijaklijent_data.sesija is not None:
        db_sesija = database.query(Sesija).filter(Sesija.id == sesijaklijent_data.sesija).first()
        if not db_sesija:
            raise HTTPException(status_code=400, detail="Sesija not found")
    else:
        raise HTTPException(status_code=400, detail="Sesija ID is required")
    sk_kwargs = dict(klijent_id=sesijaklijent_data.klijent, sesija_id=sesijaklijent_data.sesija)
    if sesijaklijent_data.id is not None:
        sk_kwargs["id"] = sesijaklijent_data.id
    db_sesijaklijent = SesijaKlijent(**sk_kwargs)
    database.add(db_sesijaklijent)
    database.commit()
    database.refresh(db_sesijaklijent)
    return db_sesijaklijent

@app.post("/sesijaklijent/bulk/", response_model=None, tags=["SesijaKlijent"])
async def bulk_create_sesijaklijent(items: list[SesijaKlijentCreate], database: Session = Depends(get_db)) -> dict:
    created_items = []
    errors = []
    for idx, item_data in enumerate(items):
        try:
            if not item_data.klijent:
                raise ValueError("Klijent ID is required")
            if not item_data.sesija:
                raise ValueError("Sesija ID is required")
            db_sesijaklijent = SesijaKlijent(id=item_data.id, klijent_id=item_data.klijent, sesija_id=item_data.sesija)
            database.add(db_sesijaklijent)
            database.flush()
            created_items.append(db_sesijaklijent.id)
        except Exception as e:
            errors.append({"index": idx, "error": str(e)})
    if errors:
        database.rollback()
        raise HTTPException(status_code=400, detail={"message": "Bulk creation failed", "errors": errors})
    database.commit()
    return {"created_count": len(created_items), "created_ids": created_items, "message": f"Successfully created {len(created_items)} SesijaKlijent entities"}

@app.delete("/sesijaklijent/bulk/", response_model=None, tags=["SesijaKlijent"])
async def bulk_delete_sesijaklijent(ids: list[int], database: Session = Depends(get_db)) -> dict:
    deleted_count = 0
    not_found = []
    for item_id in ids:
        db_sesijaklijent = database.query(SesijaKlijent).filter(SesijaKlijent.id == item_id).first()
        if db_sesijaklijent:
            database.delete(db_sesijaklijent)
            deleted_count += 1
        else:
            not_found.append(item_id)
    database.commit()
    return {"deleted_count": deleted_count, "not_found": not_found, "message": f"Successfully deleted {deleted_count} SesijaKlijent entities"}

@app.put("/sesijaklijent/{sesijaklijent_id}/", response_model=None, tags=["SesijaKlijent"])
async def update_sesijaklijent(sesijaklijent_id: int, sesijaklijent_data: SesijaKlijentCreate, database: Session = Depends(get_db)) -> SesijaKlijent:
    db_sesijaklijent = database.query(SesijaKlijent).filter(SesijaKlijent.id == sesijaklijent_id).first()
    if db_sesijaklijent is None:
        raise HTTPException(status_code=404, detail="SesijaKlijent not found")
    setattr(db_sesijaklijent, 'id', sesijaklijent_data.id)
    if sesijaklijent_data.klijent is not None:
        db_klijent = database.query(Klijent).filter(Klijent.id == sesijaklijent_data.klijent).first()
        if not db_klijent:
            raise HTTPException(status_code=400, detail="Klijent not found")
        setattr(db_sesijaklijent, 'klijent_id', sesijaklijent_data.klijent)
    if sesijaklijent_data.sesija is not None:
        db_sesija = database.query(Sesija).filter(Sesija.id == sesijaklijent_data.sesija).first()
        if not db_sesija:
            raise HTTPException(status_code=400, detail="Sesija not found")
        setattr(db_sesijaklijent, 'sesija_id', sesijaklijent_data.sesija)
    database.commit()
    database.refresh(db_sesijaklijent)
    return db_sesijaklijent

@app.delete("/sesijaklijent/{sesijaklijent_id}/", response_model=None, tags=["SesijaKlijent"])
async def delete_sesijaklijent(sesijaklijent_id: int, database: Session = Depends(get_db)):
    db_sesijaklijent = database.query(SesijaKlijent).filter(SesijaKlijent.id == sesijaklijent_id).first()
    if db_sesijaklijent is None:
        raise HTTPException(status_code=404, detail="SesijaKlijent not found")
    database.delete(db_sesijaklijent)
    database.commit()
    return {"message": "Deleted", "id": sesijaklijent_id}


############################################
#
#   GrupaKlijent functions
#
############################################

@app.get("/grupaklijent/", response_model=None, tags=["GrupaKlijent"])
def get_all_grupaklijent(detailed: bool = False, database: Session = Depends(get_db)) -> list:
    from sqlalchemy.orm import joinedload
    if detailed:
        query = database.query(GrupaKlijent)
        query = query.options(joinedload(GrupaKlijent.grupa))
        query = query.options(joinedload(GrupaKlijent.klijent))
        gk_list = query.all()
        result = []
        for gk_item in gk_list:
            item_dict = gk_item.__dict__.copy()
            item_dict.pop('_sa_instance_state', None)
            if gk_item.grupa:
                related_dict = gk_item.grupa.__dict__.copy()
                related_dict.pop('_sa_instance_state', None)
                item_dict['grupa'] = related_dict
            else:
                item_dict['grupa'] = None
            if gk_item.klijent:
                related_dict = gk_item.klijent.__dict__.copy()
                related_dict.pop('_sa_instance_state', None)
                item_dict['klijent'] = related_dict
            else:
                item_dict['klijent'] = None
            result.append(item_dict)
        return result
    else:
        return database.query(GrupaKlijent).all()

@app.get("/grupaklijent/by-grupa/{grupa_id}/", response_model=None, tags=["GrupaKlijent"])
def get_grupaklijent_by_grupa(grupa_id: int, database: Session = Depends(get_db)) -> list:
    gk_list = database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == grupa_id).all()
    result = []
    for gk in gk_list:
        item = gk.__dict__.copy()
        item.pop('_sa_instance_state', None)
        klijent = database.query(Klijent).filter(Klijent.id == gk.klijent_id).first()
        if klijent:
            klijent_dict = klijent.__dict__.copy()
            klijent_dict.pop('_sa_instance_state', None)
            item['klijent'] = klijent_dict
        result.append(item)
    return result

@app.post("/grupaklijent/", response_model=None, tags=["GrupaKlijent"])
async def create_grupaklijent(data: GrupaKlijentCreate, database: Session = Depends(get_db)):
    if data.grupa_id is None:
        raise HTTPException(status_code=400, detail="Grupa ID is required")
    if data.klijent_id is None:
        raise HTTPException(status_code=400, detail="Klijent ID is required")
    db_grupa = database.query(Grupa).filter(Grupa.id == data.grupa_id).first()
    if not db_grupa:
        raise HTTPException(status_code=400, detail="Grupa not found")
    db_klijent = database.query(Klijent).filter(Klijent.id == data.klijent_id).first()
    if not db_klijent:
        raise HTTPException(status_code=400, detail="Klijent not found")
    existing = database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == data.grupa_id, GrupaKlijent.klijent_id == data.klijent_id).first()
    if existing:
        raise HTTPException(status_code=409, detail="Klijent is already in this group")
    db_gk = GrupaKlijent(grupa_id=data.grupa_id, klijent_id=data.klijent_id)
    database.add(db_gk)
    database.commit()
    database.refresh(db_gk)
    return db_gk

@app.delete("/grupaklijent/{grupaklijent_id}/", response_model=None, tags=["GrupaKlijent"])
async def delete_grupaklijent(grupaklijent_id: int, database: Session = Depends(get_db)):
    db_gk = database.query(GrupaKlijent).filter(GrupaKlijent.id == grupaklijent_id).first()
    if db_gk is None:
        raise HTTPException(status_code=404, detail="GrupaKlijent not found")
    database.delete(db_gk)
    database.commit()
    return {"message": "Deleted", "id": grupaklijent_id}

@app.put("/grupaklijent/sync/{grupa_id}/", response_model=None, tags=["GrupaKlijent"])
async def sync_grupa_members(grupa_id: int, klijent_ids: list[int] = Body(...), database: Session = Depends(get_db)):
    db_grupa = database.query(Grupa).filter(Grupa.id == grupa_id).first()
    if not db_grupa:
        raise HTTPException(status_code=404, detail="Grupa not found")
    database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == grupa_id).delete()
    for klijent_id in klijent_ids:
        db_klijent = database.query(Klijent).filter(Klijent.id == klijent_id).first()
        if not db_klijent:
            database.rollback()
            raise HTTPException(status_code=400, detail=f"Klijent with id {klijent_id} not found")
        db_gk = GrupaKlijent(grupa_id=grupa_id, klijent_id=klijent_id)
        database.add(db_gk)
    database.commit()
    members = database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == grupa_id).all()
    return {"grupa_id": grupa_id, "member_count": len(members), "klijent_ids": [m.klijent_id for m in members]}


############################################
#
#   Sesija functions
#   ROUTE ORDER MATTERS! export-excel and mark-paid
#   MUST come BEFORE {sesija_id} routes
#
############################################

@app.get("/sesija/", response_model=None, tags=["Sesija"])
def get_all_sesija(detailed: bool = False, database: Session = Depends(get_db)) -> list:
    from sqlalchemy.orm import joinedload

    # Load payment status for ALL branches
    all_cene_paid = database.query(Cena).filter(Cena.status == "placeno").all()
    paid_sesija_ids = set()
    for c in all_cene_paid:
        if c.sesija_2_id:
            paid_sesija_ids.add(c.sesija_2_id)

    if detailed:
        query = database.query(Sesija)
        sesija_list = query.all()
        result = []
        for sesija_item in sesija_list:
            item_dict = sesija_item.__dict__.copy()
            item_dict.pop('_sa_instance_state', None)
            cena_list = database.query(Cena).filter(Cena.sesija_2_id == sesija_item.id).all()
            item_dict['cena'] = []
            for cena_obj in cena_list:
                cena_dict = cena_obj.__dict__.copy()
                cena_dict.pop('_sa_instance_state', None)
                item_dict['cena'].append(cena_dict)
            sesijaklijent_list = database.query(SesijaKlijent).filter(SesijaKlijent.sesija_id == sesija_item.id).all()
            item_dict['sesijaklijent_1'] = []
            for sesijaklijent_obj in sesijaklijent_list:
                sesijaklijent_dict = sesijaklijent_obj.__dict__.copy()
                sesijaklijent_dict.pop('_sa_instance_state', None)
                item_dict['sesijaklijent_1'].append(sesijaklijent_dict)
            sesijagrupa_list = database.query(SesijaGrupa).filter(SesijaGrupa.sesija_1_id == sesija_item.id).all()
            item_dict['sesijagrupa_1'] = []
            for sesijagrupa_obj in sesijagrupa_list:
                sesijagrupa_dict = sesijagrupa_obj.__dict__.copy()
                sesijagrupa_dict.pop('_sa_instance_state', None)
                item_dict['sesijagrupa_1'].append(sesijagrupa_dict)
            item_dict['placeno'] = sesija_item.id in paid_sesija_ids
            result.append(item_dict)
        return result
    else:
        sesija_list = database.query(Sesija).all()
        all_sk = database.query(SesijaKlijent).all()
        all_sg = database.query(SesijaGrupa).all()
        all_klijenti = database.query(Klijent).all()
        all_grupe = database.query(Grupa).all()
        klijent_map = {k.id: k for k in all_klijenti}
        grupa_map = {g.id: g for g in all_grupe}
        sk_map = {}
        for sk in all_sk:
            sk_map[sk.sesija_id] = sk.klijent_id
        sg_map = {}
        for sg in all_sg:
            sg_map[sg.sesija_1_id] = sg.grupa_id
        result = []
        for s in sesija_list:
            item = s.__dict__.copy()
            item.pop('_sa_instance_state', None)
            klijent_id = sk_map.get(s.id)
            if klijent_id:
                klijent = klijent_map.get(klijent_id)
                item['klijent_ime'] = f"{klijent.ime} {klijent.prezime}" if klijent else ""
            else:
                item['klijent_ime'] = ""
            grupa_id = sg_map.get(s.id)
            if grupa_id:
                grupa = grupa_map.get(grupa_id)
                if grupa:
                    item['grupa_naziv'] = grupa.naziv
                    item['grupa_id'] = grupa.id
                    if not item['klijent_ime']:
                        item['klijent_ime'] = f"[Grupa] {grupa.naziv}"
                else:
                    item['grupa_naziv'] = ""
                    item['grupa_id'] = None
            else:
                item['grupa_naziv'] = ""
                item['grupa_id'] = None
            item['placeno'] = s.id in paid_sesija_ids
            result.append(item)
        return result

@app.get("/sesija/count/", response_model=None, tags=["Sesija"])
def get_count_sesija(database: Session = Depends(get_db)) -> dict:
    return {"count": database.query(Sesija).count()}

@app.get("/sesija/paginated/", response_model=None, tags=["Sesija"])
def get_paginated_sesija(skip: int = 0, limit: int = 100, detailed: bool = False, database: Session = Depends(get_db)) -> dict:
    total = database.query(Sesija).count()
    sesija_list = database.query(Sesija).offset(skip).limit(limit).all()
    if not detailed:
        return {"total": total, "skip": skip, "limit": limit, "data": sesija_list}
    result = []
    for sesija_item in sesija_list:
        cena_ids = database.query(Cena.id).filter(Cena.sesija_2_id == sesija_item.id).all()
        sesijaklijent_1_ids = database.query(SesijaKlijent.id).filter(SesijaKlijent.sesija_id == sesija_item.id).all()
        sesijagrupa_1_ids = database.query(SesijaGrupa.id).filter(SesijaGrupa.sesija_1_id == sesija_item.id).all()
        item_data = {"sesija": sesija_item, "cena_ids": [x[0] for x in cena_ids], "sesijaklijent_1_ids": [x[0] for x in sesijaklijent_1_ids], "sesijagrupa_1_ids": [x[0] for x in sesijagrupa_1_ids]}
        result.append(item_data)
    return {"total": total, "skip": skip, "limit": limit, "data": result}

@app.get("/sesija/search/", response_model=None, tags=["Sesija"])
def search_sesija(database: Session = Depends(get_db)) -> list:
    return database.query(Sesija).all()

# ============================================
# REPLACE your existing export_sesija_excel function with this one
# (keep the same @app.get decorator and position in main_api.py)
# ============================================

@app.get("/sesija/export-excel/", response_model=None, tags=["Sesija"])
def export_sesija_excel(database: Session = Depends(get_db)):
    from datetime import date as date_type

    # Fetch all data
    sesija_list = database.query(Sesija).all()
    all_sk = database.query(SesijaKlijent).all()
    all_sg = database.query(SesijaGrupa).all()
    all_klijenti = database.query(Klijent).all()
    all_grupe = database.query(Grupa).all()
    all_cene = database.query(Cena).all()
    all_gk = database.query(GrupaKlijent).all()

    klijent_map = {k.id: k for k in all_klijenti}
    grupa_map = {g.id: g for g in all_grupe}
    sk_map = {sk.sesija_id: sk.klijent_id for sk in all_sk}
    sg_map = {sg.sesija_1_id: sg.grupa_id for sg in all_sg}

    # Group members map: grupa_id -> list of klijent names
    grupa_members_map = {}
    for gk in all_gk:
        if gk.grupa_id not in grupa_members_map:
            grupa_members_map[gk.grupa_id] = []
        k = klijent_map.get(gk.klijent_id)
        if k:
            grupa_members_map[gk.grupa_id].append(f"{k.ime} {k.prezime}")

    # Payment lookup
    payment_map = {}
    for c in all_cene:
        if c.status == "placeno" and c.sesija_2_id:
            if c.sesija_2_id not in payment_map:
                payment_map[c.sesija_2_id] = []
            payment_map[c.sesija_2_id].append(c)

    # Build enriched session list
    enriched = []
    for s in sesija_list:
        klijent_id = sk_map.get(s.id)
        grupa_id = sg_map.get(s.id)
        if klijent_id:
            klijent = klijent_map.get(klijent_id)
            ime = f"{klijent.ime} {klijent.prezime}" if klijent else "—"
            tip = "Individualna"
            grupa_naziv = "—"
        elif grupa_id:
            grupa = grupa_map.get(grupa_id)
            ime = f"{grupa.naziv}" if grupa else "—"
            tip = "Grupna"
            grupa_naziv = grupa.naziv if grupa else "—"
        else:
            ime = "—"
            tip = "—"
            grupa_naziv = "—"

        payments = payment_map.get(s.id, [])
        is_paid = len(payments) > 0
        nacin = payments[0].nacin_placanja if payments else "—"
        datum_uplate_val = "—"
        if payments and payments[0].datum_uplate:
            try:
                datum_uplate_val = payments[0].datum_uplate.strftime("%d.%m.%Y")
            except Exception:
                datum_uplate_val = str(payments[0].datum_uplate)

        session_cena = s.cena if s.cena else 0
        pocetak_datum = "—"
        pocetak_vreme = "—"
        kraj_vreme = "—"
        try:
            if s.pocetak:
                pocetak_datum = s.pocetak.strftime("%d.%m.%Y")
                pocetak_vreme = s.pocetak.strftime("%H:%M")
            if s.kraj:
                kraj_vreme = s.kraj.strftime("%H:%M")
        except Exception:
            pass

        # For group sessions, list members
        clanovi = ""
        if grupa_id and grupa_id in grupa_members_map:
            clanovi = ", ".join(grupa_members_map[grupa_id])

        enriched.append({
            "ime": ime,
            "tip": tip,
            "grupa_naziv": grupa_naziv,
            "clanovi": clanovi,
            "datum": pocetak_datum,
            "pocetak": pocetak_vreme,
            "kraj": kraj_vreme,
            "cena": session_cena,
            "status": (s.status or "—").capitalize(),
            "placeno": is_paid,
            "nacin": (nacin or "—").capitalize() if nacin != "—" else "—",
            "datum_uplate": datum_uplate_val,
            "sesija_id": s.id,
            "klijent_id": klijent_id,
            "grupa_id": grupa_id,
        })

    # ===== STYLES =====
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    paid_fill = PatternFill("solid", fgColor="DCFCE7")
    unpaid_fill = PatternFill("solid", fgColor="FEF3C7")
    paid_row_fill = PatternFill("solid", fgColor="F0FDF4")
    unpaid_row_fill = PatternFill("solid", fgColor="FFFBEB")
    total_font = Font(name="Arial", bold=True, size=11)
    total_fill = PatternFill("solid", fgColor="EEF2FF")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    wb = openpyxl.Workbook()

    # ============================================================
    # HELPER: write a session table to a worksheet
    # ============================================================
    def write_session_table(ws, title_text, sessions, show_payment_cols=True, show_group_members=False):
        # Determine columns
        headers = ["R.br.", "Klijent / Grupa", "Tip"]
        if show_group_members:
            headers.append("Clanovi grupe")
        headers += ["Datum", "Pocetak", "Kraj", "Cena (RSD)", "Status"]
        if show_payment_cols:
            headers += ["Placeno", "Nacin placanja", "Datum uplate"]

        # Title
        last_col_letter = get_column_letter(len(headers))
        ws.merge_cells(f"A1:{last_col_letter}1")
        title_cell = ws["A1"]
        title_cell.value = title_text
        title_cell.font = Font(name="Arial", bold=True, size=14, color="1E293B")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # Headers row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[3].height = 28

        # Column widths
        base_widths = {"R.br.": 6, "Klijent / Grupa": 25, "Tip": 14, "Clanovi grupe": 35,
                       "Datum": 14, "Pocetak": 10, "Kraj": 10, "Cena (RSD)": 14,
                       "Status": 14, "Placeno": 12, "Nacin placanja": 16, "Datum uplate": 14}
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = base_widths.get(header, 14)

        # Data rows
        row_num = 4
        total_cena = 0

        for idx, s in enumerate(sessions, 1):
            row_data = [idx, s["ime"], s["tip"]]
            if show_group_members:
                row_data.append(s["clanovi"])
            row_data += [s["datum"], s["pocetak"], s["kraj"], s["cena"], s["status"]]
            if show_payment_cols:
                row_data += ["Da" if s["placeno"] else "Ne", s["nacin"], s["datum_uplate"]]

            total_cena += s["cena"]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = cell_font
                cell.border = thin_border
                # Left-align name and members columns
                h = headers[col_idx - 1]
                cell.alignment = Alignment(
                    horizontal="left" if h in ("Klijent / Grupa", "Clanovi grupe") else "center",
                    vertical="center"
                )

            # Row coloring
            if show_payment_cols:
                placeno_col = headers.index("Placeno") + 1
                payment_cell = ws.cell(row=row_num, column=placeno_col)
                if s["placeno"]:
                    payment_cell.fill = paid_fill
                    payment_cell.font = Font(name="Arial", size=10, bold=True, color="166534")
                    for ci in range(1, len(headers) + 1):
                        if ci != placeno_col:
                            ws.cell(row=row_num, column=ci).fill = paid_row_fill
                else:
                    payment_cell.fill = unpaid_fill
                    payment_cell.font = Font(name="Arial", size=10, bold=True, color="92400E")
                    for ci in range(1, len(headers) + 1):
                        if ci != placeno_col:
                            ws.cell(row=row_num, column=ci).fill = unpaid_row_fill
            else:
                # For paid-only or unpaid-only sheets, color entire row
                is_paid_sheet = any(s["placeno"] for s in sessions[:1])  # check first item
                row_fill = paid_row_fill if sessions and sessions[0]["placeno"] else unpaid_row_fill
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=ci).fill = row_fill

            row_num += 1

        # Summary
        cena_col = headers.index("Cena (RSD)") + 1
        row_num += 1
        summary_end = cena_col - 1
        ws.merge_cells(f"A{row_num}:{get_column_letter(summary_end)}{row_num}")
        ws.cell(row=row_num, column=1, value="UKUPNO").font = total_font
        ws.cell(row=row_num, column=1).fill = total_fill
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=cena_col, value=total_cena).font = total_font
        ws.cell(row=row_num, column=cena_col).fill = total_fill
        ws.cell(row=row_num, column=cena_col).number_format = '#,##0'

        row_num += 1
        ws.merge_cells(f"A{row_num}:{get_column_letter(summary_end)}{row_num}")
        ws.cell(row=row_num, column=1, value="Broj sesija").font = Font(name="Arial", size=10)
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=cena_col, value=len(sessions)).font = Font(name="Arial", bold=True, size=10)

        return total_cena

    # ============================================================
    # SHEET 1: Sve sesije (overview)
    # ============================================================
    ws1 = wb.active
    ws1.title = "Sve sesije"
    paid_sessions = [s for s in enriched if s["placeno"]]
    unpaid_sessions = [s for s in enriched if not s["placeno"]]
    total_all = sum(s["cena"] for s in enriched)
    total_paid = sum(s["cena"] for s in paid_sessions)
    total_unpaid = sum(s["cena"] for s in unpaid_sessions)

    write_session_table(ws1, f"Izvestaj sesija — {date_type.today().strftime('%d.%m.%Y')}", enriched, show_payment_cols=True, show_group_members=True)

    # Add paid/unpaid breakdown after the summary
    last_row = ws1.max_row
    r = last_row + 1
    cena_col = 8  # adjusted for the extra "Clanovi grupe" column
    ws1.merge_cells(f"A{r}:G{r}")
    ws1.cell(row=r, column=1, value="Ukupno placeno").font = Font(name="Arial", bold=True, size=10, color="166534")
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal="right")
    ws1.cell(row=r, column=cena_col, value=total_paid).font = Font(name="Arial", bold=True, size=10, color="166534")
    ws1.cell(row=r, column=cena_col).fill = paid_fill
    ws1.cell(row=r, column=cena_col).number_format = '#,##0'

    r += 1
    ws1.merge_cells(f"A{r}:G{r}")
    ws1.cell(row=r, column=1, value="Ukupno neplaceno").font = Font(name="Arial", bold=True, size=10, color="92400E")
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal="right")
    ws1.cell(row=r, column=cena_col, value=total_unpaid).font = Font(name="Arial", bold=True, size=10, color="92400E")
    ws1.cell(row=r, column=cena_col).fill = unpaid_fill
    ws1.cell(row=r, column=cena_col).number_format = '#,##0'

    # ============================================================
    # SHEET 2: Placeno (only paid sessions with details)
    # ============================================================
    ws2 = wb.create_sheet("Placeno")
    write_session_table(ws2, f"Placene sesije — {date_type.today().strftime('%d.%m.%Y')}", paid_sessions, show_payment_cols=True, show_group_members=True)

    # ============================================================
    # SHEET 3: Neplaceno (only unpaid sessions)
    # ============================================================
    ws3 = wb.create_sheet("Neplaceno")
    write_session_table(ws3, f"Neplacene sesije — {date_type.today().strftime('%d.%m.%Y')}", unpaid_sessions, show_payment_cols=False, show_group_members=True)

    # ============================================================
    # SHEET 4: Po klijentu (per-client/group breakdown)
    # ============================================================
    ws4 = wb.create_sheet("Po klijentu")
    ws4.merge_cells("A1:G1")
    ws4["A1"].value = "Statistika po klijentu / grupi"
    ws4["A1"].font = Font(name="Arial", bold=True, size=14, color="1E293B")
    ws4["A1"].alignment = Alignment(horizontal="center")
    ws4.row_dimensions[1].height = 35

    client_headers = ["Klijent / Grupa", "Tip", "Br. sesija", "Ukupno (RSD)", "Placeno (RSD)", "Neplaceno (RSD)", "% naplaceno"]
    for col_idx, h in enumerate(client_headers, 1):
        cell = ws4.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 12
    ws4.column_dimensions["D"].width = 16
    ws4.column_dimensions["E"].width = 16
    ws4.column_dimensions["F"].width = 16
    ws4.column_dimensions["G"].width = 14

    # Aggregate per client/group
    client_stats = {}
    for s in enriched:
        key = s["ime"]
        tip = s["tip"]
        if key not in client_stats:
            client_stats[key] = {"tip": tip, "sessions": 0, "total": 0, "paid": 0, "unpaid": 0}
        client_stats[key]["sessions"] += 1
        client_stats[key]["total"] += s["cena"]
        if s["placeno"]:
            client_stats[key]["paid"] += s["cena"]
        else:
            client_stats[key]["unpaid"] += s["cena"]

    row4 = 4
    grand_total = 0
    grand_paid = 0
    grand_unpaid = 0

    for name, stats in sorted(client_stats.items()):
        ws4.cell(row=row4, column=1, value=name).font = cell_font
        ws4.cell(row=row4, column=2, value=stats["tip"]).font = cell_font
        ws4.cell(row=row4, column=2).alignment = Alignment(horizontal="center")
        ws4.cell(row=row4, column=3, value=stats["sessions"]).font = cell_font
        ws4.cell(row=row4, column=3).alignment = Alignment(horizontal="center")
        ws4.cell(row=row4, column=4, value=stats["total"]).font = cell_font
        ws4.cell(row=row4, column=4).number_format = '#,##0'
        ws4.cell(row=row4, column=5, value=stats["paid"]).font = Font(name="Arial", size=10, color="166534")
        ws4.cell(row=row4, column=5).number_format = '#,##0'
        if stats["paid"] > 0:
            ws4.cell(row=row4, column=5).fill = paid_fill
        ws4.cell(row=row4, column=6, value=stats["unpaid"]).font = Font(name="Arial", size=10, color="92400E")
        ws4.cell(row=row4, column=6).number_format = '#,##0'
        if stats["unpaid"] > 0:
            ws4.cell(row=row4, column=6).fill = unpaid_fill

        # Percentage
        pct = (stats["paid"] / stats["total"] * 100) if stats["total"] > 0 else 0
        pct_cell = ws4.cell(row=row4, column=7, value=round(pct, 1))
        pct_cell.number_format = '0.0"%"'
        pct_cell.font = cell_font
        pct_cell.alignment = Alignment(horizontal="center")
        if pct >= 100:
            pct_cell.fill = paid_fill
            pct_cell.font = Font(name="Arial", size=10, bold=True, color="166534")
        elif pct == 0:
            pct_cell.fill = unpaid_fill
            pct_cell.font = Font(name="Arial", size=10, color="92400E")

        for c in range(1, 8):
            ws4.cell(row=row4, column=c).border = thin_border

        grand_total += stats["total"]
        grand_paid += stats["paid"]
        grand_unpaid += stats["unpaid"]
        row4 += 1

    # Totals row
    row4 += 1
    ws4.merge_cells(f"A{row4}:C{row4}")
    ws4.cell(row=row4, column=1, value="UKUPNO").font = total_font
    ws4.cell(row=row4, column=1).fill = total_fill
    ws4.cell(row=row4, column=1).alignment = Alignment(horizontal="right")
    ws4.cell(row=row4, column=4, value=grand_total).font = total_font
    ws4.cell(row=row4, column=4).fill = total_fill
    ws4.cell(row=row4, column=4).number_format = '#,##0'
    ws4.cell(row=row4, column=5, value=grand_paid).font = Font(name="Arial", bold=True, size=10, color="166534")
    ws4.cell(row=row4, column=5).fill = paid_fill
    ws4.cell(row=row4, column=5).number_format = '#,##0'
    ws4.cell(row=row4, column=6, value=grand_unpaid).font = Font(name="Arial", bold=True, size=10, color="92400E")
    ws4.cell(row=row4, column=6).fill = unpaid_fill
    ws4.cell(row=row4, column=6).number_format = '#,##0'
    grand_pct = (grand_paid / grand_total * 100) if grand_total > 0 else 0
    ws4.cell(row=row4, column=7, value=round(grand_pct, 1)).font = total_font
    ws4.cell(row=row4, column=7).number_format = '0.0"%"'
    for c in range(1, 8):
        ws4.cell(row=row4, column=c).border = thin_border

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    today_str = date_type.today().strftime("%Y-%m-%d")
    filename = f"izvestaj_sesije_{today_str}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})
# ============================================
# GET single sesija — MUST be AFTER export-excel
# ============================================
@app.get("/sesija/{sesija_id}/", response_model=None, tags=["Sesija"])
async def get_sesija(sesija_id: int, database: Session = Depends(get_db)) -> Sesija:
    db_sesija = database.query(Sesija).filter(Sesija.id == sesija_id).first()
    if db_sesija is None:
        raise HTTPException(status_code=404, detail="Sesija not found")
    cena_ids = database.query(Cena.id).filter(Cena.sesija_2_id == db_sesija.id).all()
    sesijaklijent_1_ids = database.query(SesijaKlijent.id).filter(SesijaKlijent.sesija_id == db_sesija.id).all()
    sesijagrupa_1_ids = database.query(SesijaGrupa.id).filter(SesijaGrupa.sesija_1_id == db_sesija.id).all()
    return {"sesija": db_sesija, "cena_ids": [x[0] for x in cena_ids], "sesijaklijent_1_ids": [x[0] for x in sesijaklijent_1_ids], "sesijagrupa_1_ids": [x[0] for x in sesijagrupa_1_ids]}

@app.post("/sesija/", response_model=None, tags=["Sesija"])
async def create_sesija(sesija_data: SesijaCreate, database: Session = Depends(get_db)) -> Sesija:
    sesija_kwargs = dict(cena=sesija_data.cena, status=sesija_data.status, pocetak=sesija_data.pocetak, kraj=sesija_data.kraj)
    if sesija_data.id is not None:
        sesija_kwargs["id"] = sesija_data.id
    db_sesija = Sesija(**sesija_kwargs)
    database.add(db_sesija)
    database.commit()
    database.refresh(db_sesija)
    if sesija_data.uplate:
        for cena_id in sesija_data.uplate:
            db_cena = database.query(Cena).filter(Cena.id == cena_id).first()
            if not db_cena:
                raise HTTPException(status_code=400, detail=f"Cena with id {cena_id} not found")
        database.query(Cena).filter(Cena.id.in_(sesija_data.uplate)).update({Cena.sesija_2_id: db_sesija.id}, synchronize_session=False)
        database.commit()
    if sesija_data.sesijaklijent_1:
        for sesijaklijent_id in sesija_data.sesijaklijent_1:
            db_sesijaklijent = database.query(SesijaKlijent).filter(SesijaKlijent.id == sesijaklijent_id).first()
            if not db_sesijaklijent:
                raise HTTPException(status_code=400, detail=f"SesijaKlijent with id {sesijaklijent_id} not found")
        database.query(SesijaKlijent).filter(SesijaKlijent.id.in_(sesija_data.sesijaklijent_1)).update({SesijaKlijent.sesija_id: db_sesija.id}, synchronize_session=False)
        database.commit()
    if sesija_data.sesijagrupa_1:
        for sesijagrupa_id in sesija_data.sesijagrupa_1:
            db_sesijagrupa = database.query(SesijaGrupa).filter(SesijaGrupa.id == sesijagrupa_id).first()
            if not db_sesijagrupa:
                raise HTTPException(status_code=400, detail=f"SesijaGrupa with id {sesijagrupa_id} not found")
        database.query(SesijaGrupa).filter(SesijaGrupa.id.in_(sesija_data.sesijagrupa_1)).update({SesijaGrupa.sesija_1_id: db_sesija.id}, synchronize_session=False)
        database.commit()
    client_name = "Klijent"
    client_email = None
    is_group_session = False
    if sesija_data.klijent_id:
        klijent = database.query(Klijent).filter(Klijent.id == sesija_data.klijent_id).first()
        if klijent:
            client_name = f"{klijent.ime} {klijent.prezime}"
            client_email = klijent.email
            existing_sk = database.query(SesijaKlijent).filter(SesijaKlijent.sesija_id == db_sesija.id, SesijaKlijent.klijent_id == klijent.id).first()
            if not existing_sk:
                new_sk = SesijaKlijent(klijent_id=klijent.id, sesija_id=db_sesija.id)
                database.add(new_sk)
                database.commit()
    elif sesija_data.grupa_id:
        is_group_session = True
        grupa = database.query(Grupa).filter(Grupa.id == sesija_data.grupa_id).first()
        if grupa:
            existing_sg = database.query(SesijaGrupa).filter(SesijaGrupa.sesija_1_id == db_sesija.id, SesijaGrupa.grupa_id == grupa.id).first()
            if not existing_sg:
                new_sg = SesijaGrupa(grupa_id=grupa.id, sesija_1_id=db_sesija.id)
                database.add(new_sg)
                database.commit()
            group_members = database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == grupa.id).all()
            member_emails = []
            for gk in group_members:
                member = database.query(Klijent).filter(Klijent.id == gk.klijent_id).first()
                if member:
                    member_emails.append({"name": f"{member.ime} {member.prezime}", "email": member.email})
            try:
                send_session_email_to_group(action="created", grupa_naziv=grupa.naziv, pocetak=db_sesija.pocetak, kraj=db_sesija.kraj, cena=db_sesija.cena, client_emails=member_emails)
            except Exception as e:
                logger.error(f"Failed to send group email: {e}")
    else:
        sk = database.query(SesijaKlijent).filter(SesijaKlijent.sesija_id == db_sesija.id).first()
        if sk and sk.klijent_id:
            klijent = database.query(Klijent).filter(Klijent.id == sk.klijent_id).first()
            if klijent:
                client_name = f"{klijent.ime} {klijent.prezime}"
                client_email = klijent.email
    if not is_group_session:
        try:
            send_session_email(action="created", client_name=client_name, pocetak=db_sesija.pocetak, kraj=db_sesija.kraj, cena=db_sesija.cena, client_email=client_email)
        except Exception as e:
            logger.error(f"Failed to send email: {e}")
    cena_ids = database.query(Cena.id).filter(Cena.sesija_2_id == db_sesija.id).all()
    sesijaklijent_1_ids = database.query(SesijaKlijent.id).filter(SesijaKlijent.sesija_id == db_sesija.id).all()
    sesijagrupa_1_ids = database.query(SesijaGrupa.id).filter(SesijaGrupa.sesija_1_id == db_sesija.id).all()
    return {"sesija": db_sesija, "cena_ids": [x[0] for x in cena_ids], "sesijaklijent_1_ids": [x[0] for x in sesijaklijent_1_ids], "sesijagrupa_1_ids": [x[0] for x in sesijagrupa_1_ids]}

@app.post("/sesija/bulk/", response_model=None, tags=["Sesija"])
async def bulk_create_sesija(items: list[SesijaCreate], database: Session = Depends(get_db)) -> dict:
    created_items = []
    errors = []
    for idx, item_data in enumerate(items):
        try:
            db_sesija = Sesija(cena=item_data.cena, status=item_data.status, id=item_data.id, pocetak=item_data.pocetak, kraj=item_data.kraj)
            database.add(db_sesija)
            database.flush()
            created_items.append(db_sesija.id)
        except Exception as e:
            errors.append({"index": idx, "error": str(e)})
    if errors:
        database.rollback()
        raise HTTPException(status_code=400, detail={"message": "Bulk creation failed", "errors": errors})
    database.commit()
    return {"created_count": len(created_items), "created_ids": created_items, "message": f"Successfully created {len(created_items)} Sesija entities"}

@app.delete("/sesija/bulk/", response_model=None, tags=["Sesija"])
async def bulk_delete_sesija(ids: list[int], database: Session = Depends(get_db)) -> dict:
    deleted_count = 0
    not_found = []
    for item_id in ids:
        db_sesija = database.query(Sesija).filter(Sesija.id == item_id).first()
        if db_sesija:
            database.delete(db_sesija)
            deleted_count += 1
        else:
            not_found.append(item_id)
    database.commit()
    return {"deleted_count": deleted_count, "not_found": not_found, "message": f"Successfully deleted {deleted_count} Sesija entities"}

# ============================================
# MARK PAID
# ============================================
@app.post("/sesija/{sesija_id}/mark-paid/", response_model=None, tags=["Sesija"])
async def mark_sesija_paid(sesija_id: int, payment_data: dict = Body(...), database: Session = Depends(get_db)):
    from datetime import date as date_type
    db_sesija = database.query(Sesija).filter(Sesija.id == sesija_id).first()
    if db_sesija is None:
        raise HTTPException(status_code=404, detail="Sesija not found")
    existing_payment = database.query(Cena).filter(Cena.sesija_2_id == sesija_id, Cena.status == "placeno").first()
    if existing_payment:
        raise HTTPException(status_code=400, detail="Ova sesija je vec oznacena kao placena.")
    nacin_placanja = payment_data.get("nacin_placanja", "gotovina")
    datum_str = payment_data.get("datum_uplate")
    if datum_str:
        try:
            datum_uplate = date_type.fromisoformat(datum_str)
        except (ValueError, TypeError):
            datum_uplate = date_type.today()
    else:
        datum_uplate = date_type.today()
    sk = database.query(SesijaKlijent).filter(SesijaKlijent.sesija_id == sesija_id).first()
    sg = database.query(SesijaGrupa).filter(SesijaGrupa.sesija_1_id == sesija_id).first()
    if sk and sk.klijent_id:
        db_cena = Cena(cena=db_sesija.cena, status="placeno", nacin_placanja=nacin_placanja, datum_uplate=datum_uplate, sesija_2_id=sesija_id, klijent_1_id=sk.klijent_id)
        database.add(db_cena)
    elif sg and sg.grupa_id:
        group_members = database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == sg.grupa_id).all()
        if not group_members:
            db_cena = Cena(cena=db_sesija.cena, status="placeno", nacin_placanja=nacin_placanja, datum_uplate=datum_uplate, sesija_2_id=sesija_id, klijent_1_id=None)
            database.add(db_cena)
        else:
            for gk in group_members:
                db_cena = Cena(cena=db_sesija.cena, status="placeno", nacin_placanja=nacin_placanja, datum_uplate=datum_uplate, sesija_2_id=sesija_id, klijent_1_id=gk.klijent_id)
                database.add(db_cena)
    else:
        db_cena = Cena(cena=db_sesija.cena, status="placeno", nacin_placanja=nacin_placanja, datum_uplate=datum_uplate, sesija_2_id=sesija_id, klijent_1_id=None)
        database.add(db_cena)
    database.commit()
    return {"message": "Sesija oznacena kao placena", "sesija_id": sesija_id}

@app.put("/sesija/{sesija_id}/", response_model=None, tags=["Sesija"])
async def update_sesija(sesija_id: int, sesija_data: SesijaCreate, database: Session = Depends(get_db)) -> Sesija:
    db_sesija = database.query(Sesija).filter(Sesija.id == sesija_id).first()
    if db_sesija is None:
        raise HTTPException(status_code=404, detail="Sesija not found")
    setattr(db_sesija, 'cena', sesija_data.cena)
    setattr(db_sesija, 'status', sesija_data.status)
    setattr(db_sesija, 'id', sesija_data.id)
    setattr(db_sesija, 'pocetak', sesija_data.pocetak)
    setattr(db_sesija, 'kraj', sesija_data.kraj)
    if sesija_data.uplate is not None:
        database.query(Cena).filter(Cena.sesija_2_id == db_sesija.id).update({Cena.sesija_2_id: None}, synchronize_session=False)
        if sesija_data.uplate:
            for cena_id in sesija_data.uplate:
                db_cena = database.query(Cena).filter(Cena.id == cena_id).first()
                if not db_cena:
                    raise HTTPException(status_code=400, detail=f"Cena with id {cena_id} not found")
            database.query(Cena).filter(Cena.id.in_(sesija_data.uplate)).update({Cena.sesija_2_id: db_sesija.id}, synchronize_session=False)
    if sesija_data.sesijaklijent_1 is not None:
        database.query(SesijaKlijent).filter(SesijaKlijent.sesija_id == db_sesija.id).update({SesijaKlijent.sesija_id: None}, synchronize_session=False)
        if sesija_data.sesijaklijent_1:
            for sesijaklijent_id in sesija_data.sesijaklijent_1:
                db_sesijaklijent = database.query(SesijaKlijent).filter(SesijaKlijent.id == sesijaklijent_id).first()
                if not db_sesijaklijent:
                    raise HTTPException(status_code=400, detail=f"SesijaKlijent with id {sesijaklijent_id} not found")
            database.query(SesijaKlijent).filter(SesijaKlijent.id.in_(sesija_data.sesijaklijent_1)).update({SesijaKlijent.sesija_id: db_sesija.id}, synchronize_session=False)
    if sesija_data.sesijagrupa_1 is not None:
        database.query(SesijaGrupa).filter(SesijaGrupa.sesija_1_id == db_sesija.id).update({SesijaGrupa.sesija_1_id: None}, synchronize_session=False)
        if sesija_data.sesijagrupa_1:
            for sesijagrupa_id in sesija_data.sesijagrupa_1:
                db_sesijagrupa = database.query(SesijaGrupa).filter(SesijaGrupa.id == sesijagrupa_id).first()
                if not db_sesijagrupa:
                    raise HTTPException(status_code=400, detail=f"SesijaGrupa with id {sesijagrupa_id} not found")
            database.query(SesijaGrupa).filter(SesijaGrupa.id.in_(sesija_data.sesijagrupa_1)).update({SesijaGrupa.sesija_1_id: db_sesija.id}, synchronize_session=False)
    is_group_session = False
    if sesija_data.klijent_id:
        database.query(SesijaGrupa).filter(SesijaGrupa.sesija_1_id == db_sesija.id).delete()
        database.query(SesijaKlijent).filter(SesijaKlijent.sesija_id == db_sesija.id).delete()
        klijent = database.query(Klijent).filter(Klijent.id == sesija_data.klijent_id).first()
        if klijent:
            new_sk = SesijaKlijent(klijent_id=klijent.id, sesija_id=db_sesija.id)
            database.add(new_sk)
    elif sesija_data.grupa_id:
        is_group_session = True
        database.query(SesijaKlijent).filter(SesijaKlijent.sesija_id == db_sesija.id).delete()
        database.query(SesijaGrupa).filter(SesijaGrupa.sesija_1_id == db_sesija.id).delete()
        grupa = database.query(Grupa).filter(Grupa.id == sesija_data.grupa_id).first()
        if grupa:
            new_sg = SesijaGrupa(grupa_id=grupa.id, sesija_1_id=db_sesija.id)
            database.add(new_sg)
    database.commit()
    database.refresh(db_sesija)
    client_name = "Klijent"
    client_email = None
    if is_group_session and sesija_data.grupa_id:
        grupa = database.query(Grupa).filter(Grupa.id == sesija_data.grupa_id).first()
        if grupa:
            group_members = database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == grupa.id).all()
            member_emails = []
            for gk in group_members:
                member = database.query(Klijent).filter(Klijent.id == gk.klijent_id).first()
                if member:
                    member_emails.append({"name": f"{member.ime} {member.prezime}", "email": member.email})
            try:
                send_session_email_to_group(action="updated", grupa_naziv=grupa.naziv, pocetak=db_sesija.pocetak, kraj=db_sesija.kraj, cena=db_sesija.cena, client_emails=member_emails)
            except Exception as e:
                logger.error(f"Failed to send group email: {e}")
    else:
        if sesija_data.klijent_id:
            klijent = database.query(Klijent).filter(Klijent.id == sesija_data.klijent_id).first()
            if klijent:
                client_name = f"{klijent.ime} {klijent.prezime}"
                client_email = klijent.email
        else:
            sk = database.query(SesijaKlijent).filter(SesijaKlijent.sesija_id == db_sesija.id).first()
            if sk and sk.klijent_id:
                klijent = database.query(Klijent).filter(Klijent.id == sk.klijent_id).first()
                if klijent:
                    client_name = f"{klijent.ime} {klijent.prezime}"
                    client_email = klijent.email
        try:
            send_session_email(action="updated", client_name=client_name, pocetak=db_sesija.pocetak, kraj=db_sesija.kraj, cena=db_sesija.cena, client_email=client_email)
        except Exception as e:
            logger.error(f"Failed to send email: {e}")
    cena_ids = database.query(Cena.id).filter(Cena.sesija_2_id == db_sesija.id).all()
    sesijaklijent_1_ids = database.query(SesijaKlijent.id).filter(SesijaKlijent.sesija_id == db_sesija.id).all()
    sesijagrupa_1_ids = database.query(SesijaGrupa.id).filter(SesijaGrupa.sesija_1_id == db_sesija.id).all()
    return {"sesija": db_sesija, "cena_ids": [x[0] for x in cena_ids], "sesijaklijent_1_ids": [x[0] for x in sesijaklijent_1_ids], "sesijagrupa_1_ids": [x[0] for x in sesijagrupa_1_ids]}

@app.delete("/sesija/{sesija_id}/", response_model=None, tags=["Sesija"])
async def delete_sesija(sesija_id: int, database: Session = Depends(get_db)):
    db_sesija = database.query(Sesija).filter(Sesija.id == sesija_id).first()
    if db_sesija is None:
        raise HTTPException(status_code=404, detail="Sesija not found")
    pocetak = db_sesija.pocetak
    kraj = db_sesija.kraj
    cena = db_sesija.cena
    sg = database.query(SesijaGrupa).filter(SesijaGrupa.sesija_1_id == db_sesija.id).first()
    if sg and sg.grupa_id:
        grupa = database.query(Grupa).filter(Grupa.id == sg.grupa_id).first()
        if grupa:
            group_members = database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == grupa.id).all()
            member_emails = []
            for gk in group_members:
                member = database.query(Klijent).filter(Klijent.id == gk.klijent_id).first()
                if member:
                    member_emails.append({"name": f"{member.ime} {member.prezime}", "email": member.email})
            database.delete(db_sesija)
            database.commit()
            try:
                send_session_email_to_group(action="deleted", grupa_naziv=grupa.naziv, pocetak=pocetak, kraj=kraj, cena=cena, client_emails=member_emails)
            except Exception as e:
                logger.error(f"Failed to send group email: {e}")
            return {"message": "Deleted", "id": sesija_id}
    client_name = "Klijent"
    client_email = None
    sk = database.query(SesijaKlijent).filter(SesijaKlijent.sesija_id == db_sesija.id).first()
    if sk and sk.klijent_id:
        klijent = database.query(Klijent).filter(Klijent.id == sk.klijent_id).first()
        if klijent:
            client_name = f"{klijent.ime} {klijent.prezime}"
            client_email = klijent.email
    database.delete(db_sesija)
    database.commit()
    try:
        send_session_email(action="deleted", client_name=client_name, pocetak=pocetak, kraj=kraj, cena=cena, client_email=client_email)
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
    return {"message": "Deleted", "id": sesija_id}


############################################
#
#   Grupa functions
#
############################################

@app.get("/grupa/", response_model=None, tags=["Grupa"])
def get_all_grupa(detailed: bool = False, database: Session = Depends(get_db)) -> list:
    from sqlalchemy.orm import joinedload
    if detailed:
        query = database.query(Grupa)
        grupa_list = query.all()
        result = []
        for grupa_item in grupa_list:
            item_dict = grupa_item.__dict__.copy()
            item_dict.pop('_sa_instance_state', None)
            sesijagrupa_list = database.query(SesijaGrupa).filter(SesijaGrupa.grupa_id == grupa_item.id).all()
            item_dict['sesijagrupa'] = []
            for sesijagrupa_obj in sesijagrupa_list:
                sesijagrupa_dict = sesijagrupa_obj.__dict__.copy()
                sesijagrupa_dict.pop('_sa_instance_state', None)
                item_dict['sesijagrupa'].append(sesijagrupa_dict)
            gk_list = database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == grupa_item.id).all()
            item_dict['clanovi'] = []
            for gk in gk_list:
                klijent = database.query(Klijent).filter(Klijent.id == gk.klijent_id).first()
                if klijent:
                    klijent_dict = klijent.__dict__.copy()
                    klijent_dict.pop('_sa_instance_state', None)
                    item_dict['clanovi'].append(klijent_dict)
            result.append(item_dict)
        return result
    else:
        grupa_list = database.query(Grupa).all()
        result = []
        for g in grupa_list:
            item = g.__dict__.copy()
            item.pop('_sa_instance_state', None)
            member_count = database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == g.id).count()
            item['broj_clanova'] = member_count
            gk_list = database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == g.id).all()
            member_names = []
            for gk in gk_list:
                klijent = database.query(Klijent).filter(Klijent.id == gk.klijent_id).first()
                if klijent:
                    member_names.append(f"{klijent.ime} {klijent.prezime}")
            item['clanovi_imena'] = ", ".join(member_names) if member_names else "—"
            result.append(item)
        return result

@app.get("/grupa/count/", response_model=None, tags=["Grupa"])
def get_count_grupa(database: Session = Depends(get_db)) -> dict:
    return {"count": database.query(Grupa).count()}

@app.get("/grupa/paginated/", response_model=None, tags=["Grupa"])
def get_paginated_grupa(skip: int = 0, limit: int = 100, detailed: bool = False, database: Session = Depends(get_db)) -> dict:
    total = database.query(Grupa).count()
    grupa_list = database.query(Grupa).offset(skip).limit(limit).all()
    if not detailed:
        return {"total": total, "skip": skip, "limit": limit, "data": grupa_list}
    result = []
    for grupa_item in grupa_list:
        sesijagrupa_ids = database.query(SesijaGrupa.id).filter(SesijaGrupa.grupa_id == grupa_item.id).all()
        item_data = {"grupa": grupa_item, "sesijagrupa_ids": [x[0] for x in sesijagrupa_ids]}
        result.append(item_data)
    return {"total": total, "skip": skip, "limit": limit, "data": result}

@app.get("/grupa/search/", response_model=None, tags=["Grupa"])
def search_grupa(database: Session = Depends(get_db)) -> list:
    return database.query(Grupa).all()

@app.get("/grupa/{grupa_id}/", response_model=None, tags=["Grupa"])
async def get_grupa(grupa_id: int, database: Session = Depends(get_db)) -> Grupa:
    db_grupa = database.query(Grupa).filter(Grupa.id == grupa_id).first()
    if db_grupa is None:
        raise HTTPException(status_code=404, detail="Grupa not found")
    sesijagrupa_ids = database.query(SesijaGrupa.id).filter(SesijaGrupa.grupa_id == db_grupa.id).all()
    member_ids = database.query(GrupaKlijent.klijent_id).filter(GrupaKlijent.grupa_id == db_grupa.id).all()
    return {"grupa": db_grupa, "sesijagrupa_ids": [x[0] for x in sesijagrupa_ids], "member_ids": [x[0] for x in member_ids]}

@app.post("/grupa/", response_model=None, tags=["Grupa"])
async def create_grupa(grupa_data: GrupaCreate, database: Session = Depends(get_db)) -> Grupa:
    grupa_kwargs = dict(opis=grupa_data.opis, cena=grupa_data.cena, naziv=grupa_data.naziv)
    db_grupa = Grupa(**grupa_kwargs)
    database.add(db_grupa)
    database.commit()
    database.refresh(db_grupa)
    if grupa_data.sesijagrupa:
        for sesijagrupa_id in grupa_data.sesijagrupa:
            db_sesijagrupa = database.query(SesijaGrupa).filter(SesijaGrupa.id == sesijagrupa_id).first()
            if not db_sesijagrupa:
                raise HTTPException(status_code=400, detail=f"SesijaGrupa with id {sesijagrupa_id} not found")
        database.query(SesijaGrupa).filter(SesijaGrupa.id.in_(grupa_data.sesijagrupa)).update({SesijaGrupa.grupa_id: db_grupa.id}, synchronize_session=False)
        database.commit()
    if grupa_data.clanovi:
        for klijent_id in grupa_data.clanovi:
            db_klijent = database.query(Klijent).filter(Klijent.id == klijent_id).first()
            if not db_klijent:
                raise HTTPException(status_code=400, detail=f"Klijent with id {klijent_id} not found")
            db_gk = GrupaKlijent(grupa_id=db_grupa.id, klijent_id=klijent_id)
            database.add(db_gk)
        database.commit()
    sesijagrupa_ids = database.query(SesijaGrupa.id).filter(SesijaGrupa.grupa_id == db_grupa.id).all()
    member_ids = database.query(GrupaKlijent.klijent_id).filter(GrupaKlijent.grupa_id == db_grupa.id).all()
    return {"grupa": db_grupa, "sesijagrupa_ids": [x[0] for x in sesijagrupa_ids], "member_ids": [x[0] for x in member_ids]}

@app.post("/grupa/bulk/", response_model=None, tags=["Grupa"])
async def bulk_create_grupa(items: list[GrupaCreate], database: Session = Depends(get_db)) -> dict:
    created_items = []
    errors = []
    for idx, item_data in enumerate(items):
        try:
            db_grupa = Grupa(opis=item_data.opis, id=item_data.id, cena=item_data.cena, naziv=item_data.naziv)
            database.add(db_grupa)
            database.flush()
            created_items.append(db_grupa.id)
        except Exception as e:
            errors.append({"index": idx, "error": str(e)})
    if errors:
        database.rollback()
        raise HTTPException(status_code=400, detail={"message": "Bulk creation failed", "errors": errors})
    database.commit()
    return {"created_count": len(created_items), "created_ids": created_items, "message": f"Successfully created {len(created_items)} Grupa entities"}

@app.delete("/grupa/bulk/", response_model=None, tags=["Grupa"])
async def bulk_delete_grupa(ids: list[int], database: Session = Depends(get_db)) -> dict:
    deleted_count = 0
    not_found = []
    for item_id in ids:
        db_grupa = database.query(Grupa).filter(Grupa.id == item_id).first()
        if db_grupa:
            database.delete(db_grupa)
            deleted_count += 1
        else:
            not_found.append(item_id)
    database.commit()
    return {"deleted_count": deleted_count, "not_found": not_found, "message": f"Successfully deleted {deleted_count} Grupa entities"}

@app.put("/grupa/{grupa_id}/", response_model=None, tags=["Grupa"])
async def update_grupa(grupa_id: int, grupa_data: GrupaCreate, database: Session = Depends(get_db)) -> Grupa:
    db_grupa = database.query(Grupa).filter(Grupa.id == grupa_id).first()
    if db_grupa is None:
        raise HTTPException(status_code=404, detail="Grupa not found")
    setattr(db_grupa, 'opis', grupa_data.opis)
    setattr(db_grupa, 'cena', grupa_data.cena)
    setattr(db_grupa, 'naziv', grupa_data.naziv)
    if grupa_data.sesijagrupa is not None:
        database.query(SesijaGrupa).filter(SesijaGrupa.grupa_id == db_grupa.id).update({SesijaGrupa.grupa_id: None}, synchronize_session=False)
        if grupa_data.sesijagrupa:
            for sesijagrupa_id in grupa_data.sesijagrupa:
                db_sesijagrupa = database.query(SesijaGrupa).filter(SesijaGrupa.id == sesijagrupa_id).first()
                if not db_sesijagrupa:
                    raise HTTPException(status_code=400, detail=f"SesijaGrupa with id {sesijagrupa_id} not found")
            database.query(SesijaGrupa).filter(SesijaGrupa.id.in_(grupa_data.sesijagrupa)).update({SesijaGrupa.grupa_id: db_grupa.id}, synchronize_session=False)
    if grupa_data.clanovi is not None:
        database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == db_grupa.id).delete()
        for klijent_id in grupa_data.clanovi:
            db_klijent = database.query(Klijent).filter(Klijent.id == klijent_id).first()
            if not db_klijent:
                raise HTTPException(status_code=400, detail=f"Klijent with id {klijent_id} not found")
            db_gk = GrupaKlijent(grupa_id=db_grupa.id, klijent_id=klijent_id)
            database.add(db_gk)
    database.commit()
    database.refresh(db_grupa)
    sesijagrupa_ids = database.query(SesijaGrupa.id).filter(SesijaGrupa.grupa_id == db_grupa.id).all()
    member_ids = database.query(GrupaKlijent.klijent_id).filter(GrupaKlijent.grupa_id == db_grupa.id).all()
    return {"grupa": db_grupa, "sesijagrupa_ids": [x[0] for x in sesijagrupa_ids], "member_ids": [x[0] for x in member_ids]}

@app.delete("/grupa/{grupa_id}/", response_model=None, tags=["Grupa"])
async def delete_grupa(grupa_id: int, database: Session = Depends(get_db)):
    db_grupa = database.query(Grupa).filter(Grupa.id == grupa_id).first()
    if db_grupa is None:
        raise HTTPException(status_code=404, detail="Grupa not found")
    database.query(GrupaKlijent).filter(GrupaKlijent.grupa_id == grupa_id).delete()
    database.delete(db_grupa)
    database.commit()
    return {"message": "Deleted", "id": grupa_id}


############################################
#
#   Klijent functions
#
############################################

@app.get("/klijent/", response_model=None, tags=["Klijent"])
def get_all_klijent(detailed: bool = False, database: Session = Depends(get_db)) -> list:
    from sqlalchemy.orm import joinedload
    if detailed:
        query = database.query(Klijent)
        klijent_list = query.all()
        result = []
        for klijent_item in klijent_list:
            item_dict = klijent_item.__dict__.copy()
            item_dict.pop('_sa_instance_state', None)
            sesijaklijent_list = database.query(SesijaKlijent).filter(SesijaKlijent.klijent_id == klijent_item.id).all()
            item_dict['sesijaklijent'] = []
            for sesijaklijent_obj in sesijaklijent_list:
                sesijaklijent_dict = sesijaklijent_obj.__dict__.copy()
                sesijaklijent_dict.pop('_sa_instance_state', None)
                item_dict['sesijaklijent'].append(sesijaklijent_dict)
            cena_list = database.query(Cena).filter(Cena.klijent_1_id == klijent_item.id).all()
            item_dict['cena_1'] = []
            for cena_obj in cena_list:
                cena_dict = cena_obj.__dict__.copy()
                cena_dict.pop('_sa_instance_state', None)
                item_dict['cena_1'].append(cena_dict)
            result.append(item_dict)
        return result
    else:
        return database.query(Klijent).all()

@app.get("/klijent/count/", response_model=None, tags=["Klijent"])
def get_count_klijent(database: Session = Depends(get_db)) -> dict:
    return {"count": database.query(Klijent).count()}

@app.get("/klijent/paginated/", response_model=None, tags=["Klijent"])
def get_paginated_klijent(skip: int = 0, limit: int = 100, detailed: bool = False, database: Session = Depends(get_db)) -> dict:
    total = database.query(Klijent).count()
    klijent_list = database.query(Klijent).offset(skip).limit(limit).all()
    if not detailed:
        return {"total": total, "skip": skip, "limit": limit, "data": klijent_list}
    result = []
    for klijent_item in klijent_list:
        sesijaklijent_ids = database.query(SesijaKlijent.id).filter(SesijaKlijent.klijent_id == klijent_item.id).all()
        cena_1_ids = database.query(Cena.id).filter(Cena.klijent_1_id == klijent_item.id).all()
        item_data = {"klijent": klijent_item, "sesijaklijent_ids": [x[0] for x in sesijaklijent_ids], "cena_1_ids": [x[0] for x in cena_1_ids]}
        result.append(item_data)
    return {"total": total, "skip": skip, "limit": limit, "data": result}

@app.get("/klijent/search/", response_model=None, tags=["Klijent"])
def search_klijent(database: Session = Depends(get_db)) -> list:
    return database.query(Klijent).all()

@app.get("/klijent/{klijent_id}/", response_model=None, tags=["Klijent"])
async def get_klijent(klijent_id: int, database: Session = Depends(get_db)) -> Klijent:
    db_klijent = database.query(Klijent).filter(Klijent.id == klijent_id).first()
    if db_klijent is None:
        raise HTTPException(status_code=404, detail="Klijent not found")
    sesijaklijent_ids = database.query(SesijaKlijent.id).filter(SesijaKlijent.klijent_id == db_klijent.id).all()
    cena_1_ids = database.query(Cena.id).filter(Cena.klijent_1_id == db_klijent.id).all()
    return {"klijent": db_klijent, "sesijaklijent_ids": [x[0] for x in sesijaklijent_ids], "cena_1_ids": [x[0] for x in cena_1_ids]}

@app.post("/klijent/", response_model=None, tags=["Klijent"])
async def create_klijent(klijent_data: KlijentCreate, database: Session = Depends(get_db)) -> Klijent:
    klijent_kwargs = dict(ime=klijent_data.ime, email=klijent_data.email, prezime=klijent_data.prezime, broj_telefona=klijent_data.broj_telefona)
    db_klijent = Klijent(**klijent_kwargs)
    database.add(db_klijent)
    database.commit()
    database.refresh(db_klijent)
    if klijent_data.sesijaklijent:
        for sesijaklijent_id in klijent_data.sesijaklijent:
            db_sesijaklijent = database.query(SesijaKlijent).filter(SesijaKlijent.id == sesijaklijent_id).first()
            if not db_sesijaklijent:
                raise HTTPException(status_code=400, detail=f"SesijaKlijent with id {sesijaklijent_id} not found")
        database.query(SesijaKlijent).filter(SesijaKlijent.id.in_(klijent_data.sesijaklijent)).update({SesijaKlijent.klijent_id: db_klijent.id}, synchronize_session=False)
        database.commit()
    if klijent_data.cena_1:
        for cena_id in klijent_data.cena_1:
            db_cena = database.query(Cena).filter(Cena.id == cena_id).first()
            if not db_cena:
                raise HTTPException(status_code=400, detail=f"Cena with id {cena_id} not found")
        database.query(Cena).filter(Cena.id.in_(klijent_data.cena_1)).update({Cena.klijent_1_id: db_klijent.id}, synchronize_session=False)
        database.commit()
    sesijaklijent_ids = database.query(SesijaKlijent.id).filter(SesijaKlijent.klijent_id == db_klijent.id).all()
    cena_1_ids = database.query(Cena.id).filter(Cena.klijent_1_id == db_klijent.id).all()
    return {"klijent": db_klijent, "sesijaklijent_ids": [x[0] for x in sesijaklijent_ids], "cena_1_ids": [x[0] for x in cena_1_ids]}

@app.post("/klijent/bulk/", response_model=None, tags=["Klijent"])
async def bulk_create_klijent(items: list[KlijentCreate], database: Session = Depends(get_db)) -> dict:
    created_items = []
    errors = []
    for idx, item_data in enumerate(items):
        try:
            db_klijent = Klijent(ime=item_data.ime, email=item_data.email, id=item_data.id, prezime=item_data.prezime, broj_telefona=item_data.broj_telefona)
            database.add(db_klijent)
            database.flush()
            created_items.append(db_klijent.id)
        except Exception as e:
            errors.append({"index": idx, "error": str(e)})
    if errors:
        database.rollback()
        raise HTTPException(status_code=400, detail={"message": "Bulk creation failed", "errors": errors})
    database.commit()
    return {"created_count": len(created_items), "created_ids": created_items, "message": f"Successfully created {len(created_items)} Klijent entities"}

@app.delete("/klijent/bulk/", response_model=None, tags=["Klijent"])
async def bulk_delete_klijent(ids: list[int], database: Session = Depends(get_db)) -> dict:
    deleted_count = 0
    not_found = []
    for item_id in ids:
        db_klijent = database.query(Klijent).filter(Klijent.id == item_id).first()
        if db_klijent:
            database.delete(db_klijent)
            deleted_count += 1
        else:
            not_found.append(item_id)
    database.commit()
    return {"deleted_count": deleted_count, "not_found": not_found, "message": f"Successfully deleted {deleted_count} Klijent entities"}

@app.put("/klijent/{klijent_id}/", response_model=None, tags=["Klijent"])
async def update_klijent(klijent_id: int, klijent_data: KlijentCreate, database: Session = Depends(get_db)) -> Klijent:
    db_klijent = database.query(Klijent).filter(Klijent.id == klijent_id).first()
    if db_klijent is None:
        raise HTTPException(status_code=404, detail="Klijent not found")
    setattr(db_klijent, 'ime', klijent_data.ime)
    setattr(db_klijent, 'email', klijent_data.email)
    setattr(db_klijent, 'prezime', klijent_data.prezime)
    setattr(db_klijent, 'broj_telefona', klijent_data.broj_telefona)
    if klijent_data.sesijaklijent is not None:
        database.query(SesijaKlijent).filter(SesijaKlijent.klijent_id == db_klijent.id).update({SesijaKlijent.klijent_id: None}, synchronize_session=False)
        if klijent_data.sesijaklijent:
            for sesijaklijent_id in klijent_data.sesijaklijent:
                db_sesijaklijent = database.query(SesijaKlijent).filter(SesijaKlijent.id == sesijaklijent_id).first()
                if not db_sesijaklijent:
                    raise HTTPException(status_code=400, detail=f"SesijaKlijent with id {sesijaklijent_id} not found")
            database.query(SesijaKlijent).filter(SesijaKlijent.id.in_(klijent_data.sesijaklijent)).update({SesijaKlijent.klijent_id: db_klijent.id}, synchronize_session=False)
    if klijent_data.cena_1 is not None:
        database.query(Cena).filter(Cena.klijent_1_id == db_klijent.id).update({Cena.klijent_1_id: None}, synchronize_session=False)
        if klijent_data.cena_1:
            for cena_id in klijent_data.cena_1:
                db_cena = database.query(Cena).filter(Cena.id == cena_id).first()
                if not db_cena:
                    raise HTTPException(status_code=400, detail=f"Cena with id {cena_id} not found")
            database.query(Cena).filter(Cena.id.in_(klijent_data.cena_1)).update({Cena.klijent_1_id: db_klijent.id}, synchronize_session=False)
    database.commit()
    database.refresh(db_klijent)
    sesijaklijent_ids = database.query(SesijaKlijent.id).filter(SesijaKlijent.klijent_id == db_klijent.id).all()
    cena_1_ids = database.query(Cena.id).filter(Cena.klijent_1_id == db_klijent.id).all()
    return {"klijent": db_klijent, "sesijaklijent_ids": [x[0] for x in sesijaklijent_ids], "cena_1_ids": [x[0] for x in cena_1_ids]}

@app.delete("/klijent/{klijent_id}/", response_model=None, tags=["Klijent"])
async def delete_klijent(klijent_id: int, database: Session = Depends(get_db)):
    db_klijent = database.query(Klijent).filter(Klijent.id == klijent_id).first()
    if db_klijent is None:
        raise HTTPException(status_code=404, detail="Klijent not found")
    database.query(GrupaKlijent).filter(GrupaKlijent.klijent_id == klijent_id).delete()
    database.delete(db_klijent)
    database.commit()
    return {"message": "Deleted", "id": klijent_id}


############################################
# Maintaining the server
############################################
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)